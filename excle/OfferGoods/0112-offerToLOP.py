# encoding=utf-8
# 供应商数据处理为lop导入形式
__author__ = 'zxg'

import re
import sys
import os
from openpyxl import Workbook
from openpyxl.writer.excel import ExcelWriter

from util import CrawlDao, FileUtil

reload(sys)
sys.setdefaultencoding("utf-8")


class OfferToLOP:
    def __init__(self):
        self.dao = CrawlDao.CrawlDao("modeldatas")
        self.fileDao = FileUtil.FileDao()
        # 生成的excel
        wb = Workbook()
        self.write_ew = ExcelWriter(workbook=wb)
        self.lop_sheet = wb.create_sheet(u"导入lop数据", 0)
        self.wrong_sheet = wb.create_sheet(u"错误数据", 1)
        self.init_write_ew()
        # 保存的行数
        self.save_lop_row_num = 1
        self.save_wrong_row_num = 1

        # part表的基础信息sum_code:data
        self.part_dict = dict()
        self.init_part()

        # 配件库 oe:list(part_code)
        self.part_oe_code_dict = dict()
        self.init_part_goods_base()

    def init_part(self):
        part_sql = "select part_name,first_cat_name,second_cat_name,third_cat_name,sum_code from db_category_part where is_deleted = 'N'"
        result_array = self.dao.db.get_data(part_sql)
        for part_data in result_array:
            self.part_dict[str(part_data['sum_code'])] = part_data

    # 配件库中的oe码全取出来
    def init_part_goods_base(self):
        part_base_sql = "select oe_number,part_code from db_monkey_part_goods_base group by oe_number,part_code"
        result_array = self.dao.db.get_data(part_base_sql)
        for part_goods_data in result_array:
            oe_number = str(part_goods_data['oe_number'])
            part_code = str(part_goods_data['part_code'])

            if oe_number in self.part_oe_code_dict.keys():
                code_list = list(self.part_oe_code_dict[oe_number])
            else:
                code_list = list()

            code_list.append(part_code)
            self.part_oe_code_dict[oe_number] = code_list

    # 导出的excel第一行初始化
    def init_write_ew(self):
        first_column_list = [u'一级分类', u'二级分类', u'三级分类', u'商品名称（标准化)', u'标准编码', u'商品名称（供应商）', u'商品属性', u'商品品牌', u'商品类型',
                             u'规格型号', u'OE码', u'包装规格', u'购买单位', u'价格', u'城市']

        col_num = 1
        for first_column in first_column_list:
            self.lop_sheet.cell(row=1, column=col_num).value = first_column
            self.wrong_sheet.cell(row=1, column=col_num).value = first_column
            col_num += 1

    # 保存数据到sheet中
    def save_write_ew(self, column_list=list(), is_true=True):

        if is_true:
            sheet = self.lop_sheet
            self.save_lop_row_num += 1
            save_row_num = self.save_lop_row_num
        else:
            sheet = self.wrong_sheet
            self.save_wrong_row_num += 1
            save_row_num = self.save_wrong_row_num

        col_num = 1
        for first_column in column_list:
            sheet.cell(row=save_row_num, column=col_num).value = first_column
            col_num += 1


    # ===========根据不同供应商自定义处理方式=========================
    # 亚盛大众的获取的excel中的数据处理
    def yasheng_dz(self, table):
        nrows = table.nrows  # 行数
        ncols = table.ncols  # 列数

        print ('行数：%s ,列数：%s' % (nrows, ncols))
        # 遍历数据
        for rownum in range(4, nrows):
            row = table.row_values(rownum)

            oe_number = str(row[0]).strip().replace("（", "(").replace("）", ")").replace(".0", "").replace("-",
                                                                                                          "").replace(
                ".", "").replace(" ", "").replace("*", "")
            offer_name = str(row[1]).strip()

            qua = str(row[4]).strip()
            if oe_number == '':
                continue
            # 处理为lop导入的形式
            if '原厂' in qua:
                lop_qu = u'原厂'
                lop_brand = ''
            else:
                lop_qu = u'品牌'
                lop_brand = qua
            goods_type = u'全车件'
            format = ''
            package = u'1*1'
            unit = u'只'
            price = u'0'
            city = u'杭州'

            if oe_number not in self.part_oe_code_dict:
                column_list = [u'', u'', u'', u'', u'', offer_name, lop_qu, lop_brand, goods_type,
                               format, oe_number, package, unit, price, city]
                self.save_write_ew(column_list, False)
            else:
                code_list = list(self.part_oe_code_dict[oe_number])
                if len(code_list) == 1:
                    part_code = str(code_list[0])
                    part_data = dict(self.part_dict[part_code])
                    column_list = [part_data['first_cat_name'], part_data['second_cat_name'],
                                   part_data['third_cat_name'], part_data['part_name'], part_data['sum_code'],
                                   offer_name, lop_qu, lop_brand, goods_type,
                                   format, oe_number, package, unit, price, city]
                    self.save_write_ew(column_list, True)
                else:
                    for part_code in code_list:
                        part_data = dict(self.part_dict[str(part_code)])
                        column_list = [part_data['first_cat_name'], part_data['second_cat_name'],
                                       part_data['third_cat_name'], part_data['part_name'], part_data['sum_code'],
                                       offer_name, lop_qu, lop_brand, goods_type,
                                       format, oe_number, package, unit, price, city]
                        self.save_write_ew(column_list, False)

    def main(self, file_name, save_name):
        data = self.fileDao.open_excel(file_name)
        table = data.sheets()[0]

        self.yasheng_dz(table)

        self.write_ew.save(filename=save_name)


old_file = r'/Users/zxg/Documents/work/PythonExcle/供应商/ok-处理好的亚盛2500sku.xlsx'
save_file = r'/Users/zxg/Documents/work/PythonExcle/供应商/[LOP]0112亚盛LOP.xlsx'
offerLOP = OfferToLOP()
offerLOP.main(old_file, save_file)
