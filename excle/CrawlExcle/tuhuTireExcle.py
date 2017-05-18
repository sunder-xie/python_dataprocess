# encoding=utf-8
# 途虎养车轮胎

__author__ = 'zxg'

import sys

from openpyxl import Workbook
from openpyxl.writer.excel import ExcelWriter

reload(sys)
sys.setdefaultencoding("utf-8")

from util import CrawlDao


class CrawlExcle:
    def __init__(self, *name, **kwargs):
        self.dao = CrawlDao.CrawlDao()
        self.wb = None
        self.ew = None
        self.sheet = None

        # attr_key:column
        self.attr_key_col = dict()

    # 添加sheet 和 头
    def add_header_to_sheet(self,wb):
        self.sheet = wb.create_sheet(u"轮胎信息",0)
        self.sheet.cell(row=1, column=1).value = u'唯一标识'
        self.sheet.cell(row=1, column=2).value = u'名称'

        attr_sql = 'select attr_key from cm_goods_attr_value GROUP BY attr_key'
        attr_key_array = self.dao.db.get_data(attr_sql)
        indexs = 3
        for attr_key_data in attr_key_array:
            attr_key_name = attr_key_data['attr_key']
            self.sheet.cell(row=1, column=indexs).value = attr_key_name
            self.attr_key_col[attr_key_name] = indexs
            indexs += 1

    def add_data_to_sheet(self, wb):
        goods_sql = "select id,goods_name from cm_goods order by id"
        goods_array = self.dao.db.get_data(goods_sql)

        col_row = 2
        for goods_data in goods_array:
            goods_id = goods_data['id']
            goods_name = goods_data['goods_name']

            self.sheet.cell(row=col_row, column=1).value = goods_id
            self.sheet.cell(row=col_row, column=2).value = goods_name

            attr_value_sql = "select attr_key,attr_value from cm_goods_attr_value where goods_id = '"+str(goods_id)+"'"
            print 'attr_value_sql : %s' % attr_value_sql

            attr_value_array = self.dao.db.get_data(attr_value_sql)
            for attr_value_data in attr_value_array:
                col = self.attr_key_col[attr_value_data['attr_key']]
                self.sheet.cell(row=col_row, column=col).value = attr_value_data['attr_value']

            col_row += 1

    def main(self, fileName):
        print '====start ======'
        wb = Workbook()
        ew = ExcelWriter(workbook=wb)

        self.add_header_to_sheet(wb)

        self.add_data_to_sheet(wb)

        ew.save(filename=fileName)

        print '====end ======'



fileName = r'/Users/zxg/Desktop/temp/cn/tuhuTire.xlsx'
start = CrawlExcle()
start.main(fileName)