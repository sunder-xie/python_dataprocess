# encoding=utf-8
# 商车网二期

__author__ = 'zxg'

import sys

from openpyxl import Workbook
from openpyxl.writer.excel import ExcelWriter

reload(sys)
sys.setdefaultencoding("utf-8")

from util import CrawlDao


class CrawlExcle:
    def __init__(self, *name, **kwargs):
        # super(OfferGoods, self).__init__(*name, **kwargs)
        # test = load_workbook(filename = r'/Users/zxg/Desktop/temp/cn/test.xlsx')
        self.dao = CrawlDao.CrawlDao()
        self.file_name = r'/Users/zxg/Desktop/temp/cn/cn_NUMBER.xlsx'

        self.source = u'商车网'
        # base sheet
        self.car_sheet = None

        # 所有truck id 对应的行
        self.truck_id_dict = dict()
        # other sheet
        self.other_sheet = dict()
        self.other_sheet_row = 1
        self.other_sheet_max_col = 1

        # 参数对应的列
        self.attr_column = dict()

        # 所有参数
        self.attr_key_array_from_table = dict()


        # ===========车辆类型============
        # type id - name
        self.type_name = dict()

        select_type_sql = "select id,type_name from sc_truck_car_type"
        type_array = self.dao.db.get_data(select_type_sql)
        for type_data in type_array:
            id = type_data['id']
            type_name = type_data['type_name']
            self.type_name[id] = type_name

        # ===========品牌=========
        # car id - name
        self.car_name = dict()
        car_category_sql = "select id,car_name from sc_truck_car_category"
        car_array = self.dao.db.get_data(car_category_sql)
        for car_data in car_array:
            id = car_data['id']
            car_name = car_data['car_name']
            self.car_name[id] = car_name


    def create_heard(self, wb):
        self.attr_column = dict()
        self.other_sheet_row = 1
        self.other_sheet_max_col = 1

        # =========第一个基本sheet=============
        self.car_sheet = wb.create_sheet(u"基础",0)
        # 第一行
        self.car_sheet.cell(row=1, column=1).value = u'唯一标识'
        self.car_sheet.cell(row=1, column=2).value = u'名称'
        self.car_sheet.cell(row=1, column=3).value = u'批次'
        self.car_sheet.cell(row=1, column=4).value = u'类型'
        self.car_sheet.cell(row=1, column=5).value = u'车型品牌'
        self.car_sheet.cell(row=1, column=6).value = u'来源'
        self.car_sheet.cell(row=1, column=7).value = u'备注'

        # =========第二个 sheet=====
        self.other_sheet = wb.create_sheet(u"参数",1)
        # sheet 的第一行
        self.other_sheet.cell(row=1, column=1).value = u'唯一标识'


    # 生成其他sheet--第二个sheet
    def add_data_to_other_sheet(self, truck_id):
        self.other_sheet_row += 1

        self.other_sheet.cell(row=self.other_sheet_row, column=1).value = truck_id

        # self.other_sheet.write(self.other_sheet_row, 0, truck_id)

        attr_sql = "select attr_key_id,attr_value from sc_truck_car_attr where car_id = " + str(truck_id)
        print 'attr_sql : %s' % attr_sql
        try:
            attr_array = self.dao.db.get_data(attr_sql)
        except Exception, e:
            print e
            self.dao = CrawlDao.CrawlDao()
            attr_array = self.dao.db.get_data(attr_sql)
        for attr_data in attr_array:
            attr_key = attr_data['attr_key_id']
            attr_value = attr_data['attr_value']

            if attr_key not in self.attr_column.keys():
                self.find_attr_key(attr_key)
            column = self.attr_column[attr_key]
            self.other_sheet.cell(row=self.other_sheet_row, column=column).value = attr_value


    def find_attr_key(self,attr_key):
        if attr_key in self.attr_key_array_from_table.keys():
            name = self.attr_key_array_from_table[attr_key]
        else:
            attr_key_sql = "select id,key_name from sc_attr_key where id = '"+str(attr_key)+"'"
            name = self.dao.db.get_data(attr_key_sql)[0]['key_name']
            self.attr_key_array_from_table[attr_key] = name

        self.other_sheet_max_col += 1
        self.other_sheet.cell(row=1, column=self.other_sheet_max_col).value = name

        self.attr_column[attr_key] = self.other_sheet_max_col


    # 保存
    # def final_save(self):
    #     # 保存文件
    #     excle_file.save(file_name)

    def main(self):
        print '=============start================'
        # 10752 110884 211269 311294 411446
        min_id = '411446'
        file_index = 4
        all_truck_car_sql = "select id,car_name,batch_number,car_type_id,car_category_id,car_remark from sc_truck_car where source= '商车网' and batch_number != '' and id > "+min_id+" order by id limit 100000"
        print 'all_truck_car_sql : %s' % all_truck_car_sql
        truck_array = self.dao.db.get_data(all_truck_car_sql)
        truck_index = 1

        max_size = 50000

        wb = Workbook()
        ew = ExcelWriter(workbook=wb)
        self.create_heard(wb)
        file_name = self.file_name.replace("NUMBER", str(file_index))

        for truck_data in truck_array:
            # if truck_index > max_size:
            #     ew.save(filename=file_name)
            #     truck_index = 1
            #     wb = Workbook()
            #     ew = ExcelWriter(workbook=wb)
            #     self.create_heard(wb)
            #     file_index += 1
            #     file_name = self.file_name.replace("NUMBER", str(file_index))

            truck_index += 1
            truck_id = truck_data['id']
            # 数据存第一个sheet
            self.car_sheet.cell(row=truck_index, column=1).value = truck_id
            self.car_sheet.cell(row=truck_index, column=2).value = truck_data['car_name']
            self.car_sheet.cell(row=truck_index, column=3).value = truck_data['batch_number']
            self.car_sheet.cell(row=truck_index, column=4).value = self.type_name[truck_data['car_type_id']]
            self.car_sheet.cell(row=truck_index, column=5).value = self.car_name[truck_data['car_category_id']]
            self.car_sheet.cell(row=truck_index, column=6).value = self.source
            self.car_sheet.cell(row=truck_index, column=7).value = truck_data['car_remark']

            # self.car_sheet.write(truck_index, 0, truck_id)
            # self.car_sheet.write(truck_index, 1, truck_data['car_name'])
            # self.car_sheet.write(truck_index, 2, truck_data['batch_number'])
            # self.car_sheet.write(truck_index, 3, self.type_name[truck_data['car_type_id']])
            # self.car_sheet.write(truck_index, 4, self.car_name[truck_data['car_category_id']])
            # self.car_sheet.write(truck_index, 5, self.source)
            # self.car_sheet.write(truck_index, 6, truck_data['car_remark'])

            # 第二个sheet
            self.add_data_to_other_sheet(truck_id)

        if truck_index != 1:
            ew.save(filename=file_name)
        print 'end'


start = CrawlExcle()
start.main()
