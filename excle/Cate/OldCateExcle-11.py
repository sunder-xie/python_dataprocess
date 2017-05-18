# encoding=utf-8
# 2015.11.24 导出老的电商商品excle
__author__ = 'zxg'

import sys

from util import CrawlDao, StringUtil

from openpyxl import Workbook
from openpyxl.writer.excel import ExcelWriter

reload(sys)
sys.setdefaultencoding("utf-8")


class OldCateToExcle:
    def __init__(self):
        self.dao = CrawlDao.CrawlDao()
        self.stringUtil = StringUtil.StringUtil()
        self.sql_category_table = 'db_category'

        self.file_name = r'/Users/zxg/Desktop/old_cate.xlsx'

        self.cate_id_name = dict()
        self.cate_id_parent = dict()

        self.second_cat_list = list()
        # 其余行
        cate_sql = "select cat_id,cat_name,parent_id from db_category where is_deleted = 'N' order by cat_id"
        cate_array = self.dao.db.get_data(cate_sql)

        for cate_data in cate_array:
            cat_id = cate_data['cat_id']
            cat_name = cate_data['cat_name']
            parent_id = cate_data['parent_id']

            self.cate_id_name[cat_id] = cat_name
            self.cate_id_parent[cat_id] = parent_id
            if int(parent_id) != 0:
                self.second_cat_list.append(cat_id)

    def main(self):
        wb = Workbook()
        ew = ExcelWriter(workbook=wb)

        car_sheet = wb.create_sheet(u"新老分类", 0)
        # 第一行
        car_sheet.cell(row=1, column=1).value = u'一级分类'
        car_sheet.cell(row=1, column=2).value = u'二级分类'

        # 其余行
        row_num = 2
        for cat_id in self.second_cat_list:
            car_sheet.cell(row=row_num, column=1).value = self.cate_id_name[self.cate_id_parent[cat_id]]
            car_sheet.cell(row=row_num, column=2).value = self.cate_id_name[cat_id]
            row_num +=1
        ew.save(filename=self.file_name)


test = OldCateToExcle()
test.main()