# encoding=utf-8
# just test

__author__ = 'zxg'

import sys
import uuid
import xlwt
import re

from openpyxl import Workbook
from openpyxl.writer.excel import ExcelWriter

from xlrd import open_workbook
from xlutils.copy import copy

reload(sys)
sys.setdefaultencoding("utf-8")
# sys.path.append(os.getcwd() + "../util/")

from util import HttpUtil, CrawlDao, FileUtil, StringUtil
from util.pinying.pinyin import PinYin

dao = CrawlDao.CrawlDao("athena", "local")
fileDao = FileUtil.FileDao()
stringUtil = StringUtil.StringUtil()
httpUtil = HttpUtil.HttpUtil()

pinyinDO = PinYin()
pinyinDO.load_word()
print 1


monkey_dao = CrawlDao.CrawlDao("test", "local")

# liyang_dict:True or False
liyang_dict = dict()

# 是否是柴油h或电力车
def is_chaiyou_liyang(liyang_id):
    if liyang_id in liyang_dict.keys():
        the_result = liyang_dict[liyang_id]
        return the_result


    liyang_sql = "SELECT fuel_type from db_car_info_all where leyel_id = '"+liyang_id+"' limit 1"
    liyang_array = monkey_dao.db.get_data(liyang_sql)
    if len(liyang_array) == 0:
        print 'liyang nor in sql:%s' % liyang_id
        the_result = False
    else:
        fuel_type = str(liyang_array[0]['fuel_type'])
        if fuel_type == '柴油' or fuel_type == '电力':
            the_result =  True
        else:
            the_result = False

    liyang_dict[liyang_id] = the_result


    return the_result


need_up_id_list =list()
# 处理 火花塞 柴油车数据
def do_goods():
    print "start"
    the_index = 0

    goods_sql = "select uuId from db_monkey_commodity_goods where part_sum_code like '1301320%' and isdelete = 0"
    goods_array = monkey_dao.db.get_data(goods_sql)
    sum_num = len(goods_array)
    for goods_data in goods_array:
        the_index += 1
        print '处理到第 %s 个数据了，总共 %s 个goods 数据' % (str(the_index),str(sum_num))
        uuId = str(goods_data['uuId'])

        liyang_sql = "select id,liyang_Id from db_monkey_commodity_goods_car where isdelete = 0 and goods_uuId = '"+uuId+"'"
        liyang_array = monkey_dao.db.get_data(liyang_sql)

        for liyang_data in liyang_array:
            id = str(liyang_data['id'])
            liyang_id = str(liyang_data['liyang_Id'])
            if is_chaiyou_liyang(liyang_id):
                need_up_id_list.append(id)

    # 打印出来
    print '开始打印'
    max_num = 3000
    start_string = "SET unique_checks=0;SET foreign_key_checks=0;set autocommit=0;\n"
    final_string = "commit;\n"
    time_string = "select @now_time := now();\n"
    up_obj = open(r'/Users/zxg/Desktop/test/up.sql', 'w')

    up_obj.writelines(start_string)
    up_obj.writelines(time_string)

    up_help_list = list()
    for id in need_up_id_list:
        up_help_list.append(id)

        if len(up_help_list) > max_num:
            update_sql = "update db_monkey_commodity_goods_car set isdelete =1,gmt_modified = @now_time where id in ("+",".join(up_help_list)+") "
            up_obj.writelines(update_sql)
            up_obj.writelines(";\n")
            up_help_list = list()
    if len(up_help_list) > 0:
        update_sql = "update db_monkey_commodity_goods_car set isdelete =1,gmt_modified = @now_time where id in ("+",".join(up_help_list)+") "
        up_obj.writelines(update_sql)
        up_obj.writelines(";\n")
        up_help_list = list()

    up_obj.writelines(final_string)

    print "end"


do_goods()

'''
s生成 attr 数据

sql = "SELECT attr_name,attr_input_type,attr_values,attr_index,sort_order " \
      "FROM db_attribute_config " \
      "WHERE is_deleted = 'N' " \
      "AND attr_type = 0 "
the_array = dao.db.get_data(sql)
insert_help_list = list()
insert_relation_help_list = list()

attr_id = 0
for the_data in the_array:
    attr_id += 1
    attr_val_array = str(the_data['attr_values']).split("\n")
    for attr_val in attr_val_array:
        re_data = {"attr_id": attr_id, "attr_value": attr_val, "gmt_create": "@now_time", "gmt_modified": "@now_time"}
        insert_relation_help_list.append(re_data)

    save_data = {
        "id": attr_id,
        "attr_name": str(the_data['attr_name']),
        "sort_order": str(the_data['sort_order']),
        "gmt_create": "@now_time", "gmt_modified": "@now_time"
    }
    insert_help_list.append(save_data)

print dao.get_batch_sql("center_attr_config", insert_help_list)
print "  "
print "  "
print dao.get_batch_sql("center_attr_value_relation", insert_relation_help_list)

'''
'''
excle_file = r'/Users/zxg/Desktop/商品属性0626_合并sheet(1).xlsx'
# 单个excle处理
data = fileDao.open_excel(excle_file)
#
table = data.sheets()[0]

n_rows = table.nrows  # 行数
n_cols = table.ncols  # 列数
print ('行数：%s ,列数：%s' % (n_rows, n_cols))

wb = Workbook()
ew = ExcelWriter(workbook=wb)
the_sheet = wb.create_sheet("1", 0)
the_sheet.cell(row=1, column=1).value = u'商品编码'
the_sheet.cell(row=1, column=2).value = u'品质'

start_row_num = 2
for rownum in range(1, n_rows):
    row = table.row_values(rownum)
    goods_sn = str(row[0]).strip()
    standard = str(row[10]).strip().upper()
    origin = str(row[11]).strip().upper()

    is_need_write = False
    if "正厂" in standard or "正厂" in origin or "ZC" in standard or "ZC" in origin:
        is_need_write = True
        quality = "正厂原厂"
    elif "拆车" in standard or "拆车" in origin:
        is_need_write = True
        quality = "全新拆车"

    if is_need_write:
        the_sheet.cell(row=start_row_num, column=1).value = goods_sn
        the_sheet.cell(row=start_row_num, column=2).value = quality
        start_row_num += 1

ew.save(filename=r'/Users/zxg/Desktop/quality.xlsx')
'''

'''

# 加首字母
car_code_dict = dict()

cat_array = dao.db.get_data("select id,cat_code,cat_level,parent_id,cat_name from center_category")
for cat_data in cat_array:
    id = str(cat_data['id'])
    parent_id = str(cat_data['parent_id'])
    cat_code = str(cat_data['cat_code'])
    cat_name = str(cat_data['cat_name'])
    cat_level = str(cat_data['cat_level'])

    if cat_level == '3':
        car_code_dict[cat_code.replace(".","")] = id


    first_letter = str(pinyinDO.firstLetter(cat_name))
    update_sql = "update center_category set first_letter = '"+first_letter+"' where id = '"+id+"'"
    print update_sql
    dao.db.update_data(update_sql)

# part
part_array = dao.db.get_data("select id,sum_code from center_part where is_deleted = 'N' ")
for part_data in part_array:
    id = str(part_data['id'])
    sum_code = str(part_data['sum_code'])

    third_code = sum_code[:7]
    third_id = car_code_dict[third_code]

    update_sql = "update center_part set third_cate_id = '"+third_id+"' where id = '"+id+"'"
    print update_sql
    dao.db.update_data(update_sql)
'''
'''
# 新增
# save_data = {'part_name': '仪表台扬声器', 'part_code': '00', 'first_cat_id': '443',
#              'first_cat_name': '电器系统', 'second_cat_id': '521', 'second_cat_name': '仪表台',
#              'third_cat_id': '542', 'third_cat_name': '仪表台扬声器',
#              'sum_code': '110418000', 'cat_kind': '1', 'vehicle_code': 'C'}
# dao.insert_temple("db_category_part",
#                   save_data)
# ========================================================================================================


part_dict = dict()
part_array = dao.db.get_data("SELECT part_name,sum_code "
                             "FROM db_category_part WHERE is_deleted = 'N' AND vehicle_code BETWEEN 'C' and 'CH'")
for part_data in part_array:
    part_dict[str(part_data['part_name'])] = str(part_data['sum_code'])

h_list = list()
part_h_array = dao.db.get_data("SELECT part_name "
                             "FROM db_category_part WHERE is_deleted = 'N' AND vehicle_code = 'H'")
for one in part_h_array:
    h_list.append(str(one['part_name']))

newname = r'/Users/zxg/Documents/work/淘气档口/work/2016.06 monkey－part/相同oe不同标准零件名称（完成）.xls'
# 单个excle处理
data = fileDao.open_excel(newname)
table = data.sheets()[0]
nrows = table.nrows  # 行数
ncols = table.ncols  # 列数

name_cols_num = 0
code_cols_num = 0

rb = open_workbook(newname)

wb = copy(rb)

# 通过get_sheet()获取的sheet有write()方法
ws = wb.get_sheet(0)

# 遍历所有行，把标准零件编码进行更改
has_show_list = list()
for rownum in range(1, nrows):
    row = table.row_values(rownum)
    id = str(row[0])
    part_name = str(row[9])
    part_code = str(row[10]).replace(".0", "")
    if part_name == '':
        print 'id：%s,为空part name' % id
        continue

    if part_name not in part_dict.keys():
        if part_name not in h_list:
            print 'part name not in part_dict:%s'%part_name
            has_show_list.append(part_name)

        continue
    new_part_code = part_dict[part_name]
    if part_code != '':
        if part_code != new_part_code:
            print 'id：%s,part name 和code不一致,part name:%s' % (id,part_name)
            # has_show_list.append(part_name)
            ws.write(rownum, 10, new_part_code.decode('UTF-8'))
            continue
    else:
        ws.write(rownum, 10, new_part_code.decode('UTF-8'))
#
wb.save(newname)
'''

'''
update_list = list()
update_list.append("SET unique_checks=0;SET foreign_key_checks=0;set autocommit=0")
brand_array = dao.db.get_data("select car_brand,brand_first_word from db_car_info_all where brand_first_word is NULL group by car_brand")
for brand_data in brand_array:
    car_brand = str(brand_data['car_brand'])
    brand_first_word = pinyinDO.firstLetter(car_brand)

    update_list.append("update db_car_info_all set brand_first_word = '"+brand_first_word+"' where car_brand = '" + car_brand +"'")

update_list.append("commit;")

file_object = open(r'/Users/zxg/Documents/work/sql/every-day/2016-05-27/update_db_car_info_all.sql', 'w')

try:
    file_object.writelines(";\n".join(update_list))
finally:
    file_object.close()
'''

'''
excle = r'/Users/zxg/Documents/test/show/wwwwwwwwwwwwww/一汽奥迪-A3 Limousine-20151208.xls'
newname = excle
portion = os.path.splitext(excle)  # 如果后缀是.txt
if portion[1] == '.xlsx':
    newname = portion[0] + ".xls"
    os.rename(excle, newname)

rb = open_workbook(newname)

wb = copy(rb)

# 通过get_sheet()获取的sheet有write()方法
ws = wb.get_sheet(0)
ws.write(1, 3, "我是啊".decode('UTF-8'))
wb.save(newname)
'''

'''# E
car_dict = dict()
china_car_sql = "select id,car_name,series_name,car_code from bmw_car where produce_area = 'CHN' "
car_array = dao.db.get_data(china_car_sql)
for car_data in car_array:
    car_id = str(car_data['id'])
    car_string = str(car_data['car_code'])+"-"+str(car_data['series_name'])+"-"+str(car_data['car_name'])
    car_dict[car_id] = car_string

goods_relation_dict = dict()
relation_sql = "select bgc.goods_id as goods_id,bgc.car_id as car_id from bmw_car bc,bmw_goods_car_cat_relation bgc where bc.produce_area ='CHN' and bgc.car_id = bc.id"
relation_array = dao.db.get_data(relation_sql)
for relation_data in relation_array:
    goods_id = str(relation_data['goods_id'])
    car_id = str(relation_data['car_id'])

    car_string = car_dict[car_id]
    if goods_id in goods_relation_dict.keys():
        string_list = goods_relation_dict[goods_id]
    else:
        string_list = list()
    string_list.append(car_string)
    goods_relation_dict[goods_id] = string_list



wb = Workbook()
ew = ExcelWriter(workbook=wb)

cat_sheet = wb.create_sheet(u"1", 0)
cat_sheet.cell(row=1, column=1).value = u'goods_id'
cat_sheet.cell(row=1, column=2).value = u'oe'
cat_sheet.cell(row=1, column=3).value = u'名称'
cat_sheet.cell(row=1, column=4).value = u'开始时间'
cat_sheet.cell(row=1, column=5).value = u'结束时间'
cat_sheet.cell(row=1, column=6).value = u'车型'

row_num = 2

goods_sql = "select bg.id as id,bg.oe_number as oe_number,bg.goods_name as goods_name,bg.from_time as from_time,bg.end_time as end_time " \
            "from bmw_car bc,bmw_goods_car_cat_relation bgc,bmw_goods bg " \
            "where bc.produce_area ='CHN' " \
            "and bgc.car_id = bc.id and bgc.goods_id = bg.id"
goods_array = dao.db.get_data(goods_sql)
for goods_data in goods_array:
    goods_id = str(goods_data['id'])
    cat_sheet.cell(row=row_num, column=1).value = goods_id
    cat_sheet.cell(row=row_num, column=2).value = str(goods_data['oe_number'])
    cat_sheet.cell(row=row_num, column=3).value = str(goods_data['goods_name'])
    cat_sheet.cell(row=row_num, column=4).value = str(goods_data['from_time'])
    cat_sheet.cell(row=row_num, column=5).value = str(goods_data['end_time'])
    cat_sheet.cell(row=row_num, column=6).value = "/".join(goods_relation_dict[goods_id])
    row_num += 1

ew.save(filename=r'/Users/zxg/Documents/bmw-chn.xlsx')
'''

'''

wb = Workbook()
ew = ExcelWriter(workbook=wb)

cat_sheet = wb.create_sheet(u"1", 0)

cat_array = dao.db.get_data("select cat_name,cat_level,cat_code from db_category where is_deleted = 'N' group by cat_name,cat_level,cat_code ORDER BY CAT_ID")

row_num = 0
for cat_data in cat_array:
    row_num += 1
    cat_name = str(cat_data['cat_name'])
    cat_level = str(cat_data['cat_level'])
    cat_code = str(cat_data['cat_code'])

    true_code = cat_code
    if cat_level == "2":
        true_code = cat_code[0:2]+"."+cat_code[2:]
    if cat_level == "3":
        true_code = cat_code[0:2]+"."+cat_code[2:4]+"."+cat_code[4:]

    cat_sheet.cell(row=row_num, column=1).value = true_code
    cat_sheet.cell(row=row_num, column=2).value = cat_name

ew.save(filename=r'/Users/zxg/Documents/金蝶.xlsx')
'''

'''
cat_dict = dict()
cat_sql = "select cat_id,cat_code,parent_id from db_category where is_deleted = 'N'"
cat_array = dao.db.get_data(cat_sql)
for cat_data in cat_array:
    cat_dict[str(cat_data['cat_id'])] = cat_data


def get_code(code="",parent_id=""):
    if parent_id == '0':
        return code
    parent_data = cat_dict[parent_id]
    parent_code = str(parent_data['cat_code'])
    this_id = str(parent_data['parent_id'])

    new_code = parent_code+code
    return get_code(new_code, this_id)

# 更新每个code
update_sql_list = list()

for up_cat in cat_array:
    parent_id = str(up_cat['parent_id'])
    cat_code = str(up_cat['cat_code'])
    cat_id = str(up_cat['cat_id'])
    if parent_id == '0':
        continue
    new_code = get_code(cat_code, parent_id)
    update_sql = "update db_category set cat_code = '"+new_code+"',gmt_modified = now() where cat_id = '"+cat_id+"'"
    update_sql_list.append(update_sql)




file_object = open(r'/Users/zxg/Desktop/temp/元旦cat编码更改/cat_code_update.txt', 'w')

try:
    file_object.writelines(";\n".join(update_sql_list))
finally:
    file_object.close()
'''

'''
cat_dict = dict()
cat_sql = "select cat_name,cat_id,vehicle_code from center_category where is_deleted = 'N'"
cat_array = dao.db.get_data(cat_sql)
for cat_data in cat_array:
    cat_id = str(cat_data['cat_id'])
    cat_dict[cat_id] = cat_data

center_sql = "select cgcr.car_id as car_id,cgb.first_cat_id as first_cat_id,cgb.second_cat_id as second_cat_id,cgb.third_cat_id as third_cat_id " \
             "from center_goods_base cgb,center_goods cg,center_goods_car_relation cgcr " \
"where cgb.id = cg.goods_base_id and cg.id = cgcr.goods_id " \
"group by cgcr.car_id,cgb.first_cat_id,cgb.second_cat_id,cgb.third_cat_id "
center_array = dao.db.get_data(center_sql)

insert_list = list()
for center_data in center_array:
    first_cat_id = str(center_data['first_cat_id'])
    second_cat_id = str(center_data['second_cat_id'])
    third_cat_id = str(center_data['third_cat_id'])

    first_cat_name = cat_dict[first_cat_id]['cat_name']
    second_cat_name = cat_dict[second_cat_id]['cat_name']

    third_data = cat_dict[third_cat_id]
    third_cat_name = third_data['cat_name']
    vehicle_code = third_data['vehicle_code']

    center_data['first_cat_name'] = first_cat_name
    center_data['second_cat_name'] = second_cat_name
    center_data['third_cat_name'] = third_cat_name
    center_data['vehicle_code'] = vehicle_code
    insert_list.append(center_data)

    if len(insert_list) > 10000:
        dao.insert_batch_temple("center_car_cate_relation",insert_list)
        insert_list = list()


if len(insert_list) > 0:
    dao.insert_batch_temple("center_car_cate_relation",insert_list)
'''
'''
oe_dict = dict()
all_oe_list = set()
stable_oe_sql = "select oe_number_trim,oe_number " \
                "from center_goods cg, " \
                "( select goods_id from center_goods_car_relation " \
                "where car_brand = '五菱' " \
                "and car_series = '宏光' " \
                "group by goods_id " \
                ") a " \
                "where cg.id = a.goods_id "
for oe_data in dao.db.get_data(stable_oe_sql):
    data_oe = str(oe_data['oe_number_trim']).replace("（", "(").replace("）", ")").replace(".0", "").replace("-", "").replace(
                ".", "").replace(" ", "").replace("*", "")
    if data_oe.startswith("L") or data_oe.startswith("l"):
        data_oe = data_oe[1:]

    all_oe_list.add(data_oe)
    oe_dict[data_oe] = str(oe_data['oe_number'])


excle_file = r'/Users/zxg/Documents/work/PythonExcle/供应商/五菱宏光LZW6430-20160318.xls'

# 单个excle处理
data = fileDao.open_excel(excle_file)
#
table = data.sheets()[0]

n_rows = table.nrows  # 行数
n_cols = table.ncols  # 列数
print ('行数：%s ,列数：%s' % (n_rows, n_cols))
in_num = 0
not_in_num = 0



wb = Workbook()
ew = ExcelWriter(workbook=wb)

car_sheet = wb.create_sheet(u"1", 0)


for row_num in range(3, n_rows):
    row = table.row_values(row_num)

    old_oe_number = str(row[1]).replace("（", "(").replace("）", ")").replace(".0", "").replace("-", "").replace(
                ".", "").replace(" ", "").replace("*", "")

    if old_oe_number.startswith("L") or old_oe_number.startswith("l"):
        old_oe_number = old_oe_number[1:]
    if old_oe_number in all_oe_list:
        in_num += 1
        # print '原始：%s  去符号后：%s' %(str(row[1]),old_oe_number)
        for col_num in range(0,n_cols):
            car_sheet.cell(row=in_num, column=col_num+1).value = str(row[col_num]).replace("（", "(").replace("）", ")").replace(".0", "").replace("-", "").replace(
                ".", "").replace(" ", "").replace("*", "").strip()

        car_sheet.cell(row=in_num, column=n_cols+1).value = oe_dict[old_oe_number]
    else:
        not_in_num += 1

print "in:%s,not in:%s" % (str(in_num),str(not_in_num))
ew.save(filename=r'/Users/zxg/Documents/五菱宏光LZW6430-20160318-匹配上.xlsx')
'''

'''
cat_dict = dict()
cat_sql = "select cat_id,cat_name,cat_code,vehicle_code from db_category where is_deleted = 'N'"
for cat_data in dao.db.get_data(cat_sql):
    cat_id = int(cat_data['cat_id'])
    cat_dict[cat_id] = cat_data


wb = Workbook()
ew = ExcelWriter(workbook=wb)

cat_sheet = wb.create_sheet(u"1", 0)

first_row_list = [u'id',u'oe 码',u'标准零件名称',u'标准零件编码',u'一级分类',u'一级编码',u'二级分类',u'二级编码',u'三级名称',u'三级编码',u'车辆种类码',]
cat_sheet.cell(row=1, column=1).value = u'id'
cat_sheet.cell(row=1, column=2).value = u'oe 码'
cat_sheet.cell(row=1, column=3).value = u'标准零件名称'
cat_sheet.cell(row=1, column=4).value = u'标准零件编码'
cat_sheet.cell(row=1, column=5).value = u'一级分类'
cat_sheet.cell(row=1, column=6).value = u'一级编码'
cat_sheet.cell(row=1, column=7).value = u'二级分类'
cat_sheet.cell(row=1, column=8).value = u'二级编码'
cat_sheet.cell(row=1, column=9).value = u'三级名称'
cat_sheet.cell(row=1, column=10).value = u'三级编码'
cat_sheet.cell(row=1, column=11).value = u'车辆种类码'
cat_sheet.cell(row=1, column=12).value = u'备注'
cat_sheet.cell(row=1, column=13).value = u'适配车型'


part_sql = "select dmp.oe_number as oe_number,dmp.part_name as part_name,dmp.part_code as part_code,dmp.first_cate_id as first_cate_id,dmp.second_cate_id as second_cate_id, " \
"dmp.third_cate_id as third_cate_id ,dmp.id as id ,dpg.remarks as remarks,dpg.uuId as goods_uuid  " \
"from db_monkey_part_goods_base dmp, " \
"( " \
"select oe_number " \
"from db_monkey_part_goods_base " \
"group by oe_number " \
"having count(*) > 1 " \
") more, db_monkey_part_goods dpg " \
"where more.oe_number =  dmp.oe_number and dpg.goods_base_id = dmp.uuId"
row_num = 2
for part_data in dao.db.get_data(part_sql):
    first_cat_id = int(part_data['first_cate_id'])
    second_cate_id = int(part_data['second_cate_id'])
    third_cate_id = int(part_data['third_cate_id'])

    first_data = cat_dict[first_cat_id]
    second_data = cat_dict[second_cate_id]
    third_data = cat_dict[third_cate_id]

    cat_sheet.cell(row=row_num, column=1).value = str(part_data['id'])
    cat_sheet.cell(row=row_num, column=2).value = str(part_data['oe_number'])
    cat_sheet.cell(row=row_num, column=3).value = str(part_data['part_name'])
    cat_sheet.cell(row=row_num, column=4).value = str(part_data['part_code'])
    cat_sheet.cell(row=row_num, column=5).value = first_data['cat_name']
    cat_sheet.cell(row=row_num, column=6).value = first_data['cat_code']
    cat_sheet.cell(row=row_num, column=7).value = second_data['cat_name']
    cat_sheet.cell(row=row_num, column=8).value = second_data['cat_code']
    cat_sheet.cell(row=row_num, column=9).value = third_data['cat_name']
    cat_sheet.cell(row=row_num, column=10).value = third_data['cat_code']
    cat_sheet.cell(row=row_num, column=11).value = third_data['vehicle_code']
    cat_sheet.cell(row=row_num, column=12).value = str(part_data['remarks'])


    goods_uuid = str(part_data['goods_uuid'])
    car_string = ""
    car_sql = "select liyang_brand,liyang_factory,liyang_series,liyang_model from db_monkey_part_liyang_relation where goods_id = '"+goods_uuid+"' group by liyang_model"
    for car_data in dao.db.get_data(car_sql):
        car_string += str(car_data['liyang_brand']) + "-"+str(car_data['liyang_factory']) + "-"+str(car_data['liyang_series']) + "-"+str(car_data['liyang_model']) +" "
    cat_sheet.cell(row=row_num, column=13).value = car_string

    row_num += 1

ew.save(filename=r'/Users/zxg/Documents/oe多分类-other.xlsx')
'''
'''
all_oe_list = list()
stable_oe_sql = "SELECT oe_number FROM db_monkey_part_goods_base GROUP BY oe_number"
for oe_data in dao.db.get_data(stable_oe_sql):
    data_oe = str(oe_data['oe_number'])
    all_oe_list.append(data_oe)

wb = Workbook()
ew = ExcelWriter(workbook=wb)

cat_sheet = wb.create_sheet(u"1", 0)

cat_sheet.cell(row=1, column=1).value = u'oe 码'
cat_sheet.cell(row=1, column=2).value = u'零件名称'
cat_sheet.cell(row=1, column=3).value = u'车型'
cat_sheet.cell(row=1, column=4).value = u'产地'
cat_sheet.cell(row=1, column=5).value = u'淘汽价格'
cat_sheet.cell(row=1, column=6).value = u'商品属性'
cat_sheet.cell(row=1, column=7).value = u'淘汽oe'
cat_sheet.cell(row=1, column=8).value = u'单位'

excle_file = r'/Users/zxg/Documents/work/PythonExcle/供应商/杭州韩现数据20160106.xls'
# excle_file = r'/Users/zxg/Documents/work/PythonExcle/供应商/亚盛160106.xls'

# 单个excle处理
data = fileDao.open_excel(excle_file)
#
table = data.sheets()[0]

n_rows = table.nrows  # 行数
n_cols = table.ncols  # 列数

in_num = 0
not_in_num = 0
for row_num in range(1, n_rows):
    row = table.row_values(row_num)

    old_oe_number = str(row[0])
    attr = str(row[3]).strip()

    old_oe_list = list()
    if "/" in old_oe_number and "-" in old_oe_number:
        old_oe_array = old_oe_number.split("-")
        more_array = old_oe_array[1].split("/")
        for value in more_array:
            oe = old_oe_array[0]+"-"+value
            old_oe_list.append(oe)
    else:
        old_oe_list.append(old_oe_number)

    for now_old_oe_number in old_oe_list:
        old_oe_array = now_old_oe_number.split("-")
        if len(old_oe_array) > 2 and attr != '正厂':
            now_old_oe_number = old_oe_array[0]+old_oe_array[1]
        oe_number = now_old_oe_number.strip().replace("（", "(").replace("）", ")").replace(".0", "").replace("-", "").replace(
            ".", "").replace(" ", "").replace("*", "")

        if oe_number == '':
            not_in_num += 1
            continue
        if oe_number in all_oe_list:
            in_num += 1


            new_attr = attr
            if '原厂' in attr :
               new_attr = '正厂'
            if '配套' in attr :
               new_attr = '配套'
            if '下线' in attr :
               new_attr = '下线'
            if attr == '拆车':
               new_attr = '旧车拆车'
            if attr == '拆车新':
               new_attr = '新车拆车'

            cat_sheet.cell(row=in_num + 1, column=1).value = old_oe_number
            cat_sheet.cell(row=in_num + 1, column=2).value = str(row[1])
            cat_sheet.cell(row=in_num + 1, column=3).value = str(row[2])
            cat_sheet.cell(row=in_num + 1, column=4).value = attr
            cat_sheet.cell(row=in_num + 1, column=5).value = ''
            cat_sheet.cell(row=in_num + 1, column=6).value = new_attr
            cat_sheet.cell(row=in_num + 1, column=7).value = oe_number
            cat_sheet.cell(row=in_num + 1, column=8).value = str(row[4])
            print old_oe_number
        else:
            not_in_num += 1

print "in:%s ;not in:%s" % (str(in_num), str(not_in_num))


ew.save(filename=r'/Users/zxg/Documents/work/PythonExcle/供应商/韩现10000sku.xlsx')
'''

'''
update_sql_list = list()

first_code_dict = {'1': '10', '2': '11', '3': '12', '4': '13', '5': '14', '6': '15', 'A': '16', 'B': '17', 'C': '18',
                   'D': '19', 'Z': '20'}
# category_sql = "SELECT cat_id,cat_code FROM db_category WHERE is_deleted = 'N' AND cat_level = 1"
# for cat_data in dao.db.get_data(category_sql):
#     cat_code = str(cat_data['cat_code'])
#     old_cat_code = cat_code
#     if cat_code in first_code_dict.keys():
#         cat_code = first_code_dict[cat_code]
#     else:
#         print 'wrong'+cat_code
#
#     cat_id = str(cat_data['cat_id'])
#     update_sql = "UPDATE db_category SET cat_code = '" + str(
#         cat_code) + "',gmt_modified = now() WHERE cat_code = '" + old_cat_code + "'"
#     update_sql_list.append(update_sql)
#
#     dao.db.update_data(update_sql)
#
# update_contain_x_sql = "UPDATE db_category SET cat_code = '000',gmt_modified = now() WHERE is_deleted = 'N' AND cat_code = 'XXX'"
# update_sql_list.append(update_contain_x_sql)
#
# dao.db.update_data(update_contain_x_sql)

part_sql = "SELECT id,sum_code FROM db_category_part WHERE is_deleted = 'N'"
for part_data in dao.db.get_data(part_sql):
    sum_code = str(part_data['sum_code'])
    id = str(part_data['id'])

    first_code = sum_code[0:1]
    second_code = sum_code[1:3]
    third_code = sum_code[3:6]
    part_code = sum_code[6:]

    if first_code in first_code_dict.keys():
        first_code = str(first_code_dict[first_code])
    else:
        first_code = "0" + first_code
    if third_code == 'XXX':
        third_code = '000'
    if part_code == 'XX':
        part_code = '00'

    new_sum_code = first_code + second_code + third_code + part_code

    update_sql = "UPDATE db_category_part SET sum_code = '" + new_sum_code + "' WHERE sum_code = '" + sum_code + "'"
    update_sql_list.append(update_sql)

    dao.db.update_data(update_sql)

file_object = open(r'/Users/zxg/Desktop/temp/元旦cat编码更改/monkey_update.txt', 'w')

try:
    file_object.writelines(";\n".join(update_sql_list))
finally:
    file_object.close()
'''

'''
goods_relation_sql = "select old_cat_id from db_category_cat_relation_temp where cat_status = 1 group by old_cat_id"

for goods_data in dao.db.get_data(goods_relation_sql):

    old_cat_id = str(goods_data['old_cat_id'])

    goods_relation = "select new_third_cat_id from db_category_cat_relation_temp where old_cat_id = "+old_cat_id
    for new_data in dao.db.get_data(goods_relation):
        new_third_cat_id = str(new_data['new_third_cat_id'])

        goods_sql = "select count(1) as num from db_goods where cat_id = "+new_third_cat_id
        num = int(dao.db.get_data(goods_sql)[0]['num'])

        update_data = {
            'old_cat_id': old_cat_id,
            'new_third_cat_id': new_third_cat_id,
            'goods_num': num
        }

        exist_data = {
            'old_cat_id': old_cat_id,
            'new_third_cat_id': new_third_cat_id
        }

        dao.update_temple("db_category_cat_relation_temp",update_data,exist_data)
'''

'''# 临时处理反推老类目数据
yes_set = set()
new_set = set()
excel = fileDao.open_excel(r'/Users/zxg/Documents/tempFile/新类目反推老类目－重复数据(2).xls')
sheet = excel.sheet_by_index(0)
for row_num in range(1, sheet.nrows):
    row = sheet.row_values(row_num)
    old = str(row[3]).replace(".0", "").strip()
    new = str(row[5]).replace(".0", "").strip()
    status = str(row[10]).replace(".0", "").strip()
    new_set.add(new)

    if status == 'Y':
        yes_set.add(new)

        exist_data = {
            "old_cat_id": old,
            "new_third_cat_id": new,
            "cat_status": 2,
        }
        dao.update_temple("db_category_cat_relation_temp", {"cat_status": 0}, exist_data)

print len(new_set)
for new_id in new_set:
    if new_id not in yes_set:
        print new_id
'''

'''# 临时处理新增的goods
update_string = list()
goods_sql = "select cat_id,goods_id from db_goods where goods_id > 332358"
for goods_data in dao.db.get_data(goods_sql):
    old_cat_id = str(goods_data['cat_id'])
    goods_id = str(goods_data['goods_id'])

    cat_sql = "select old_cat_id,old_cat_name,new_third_cat_id,new_third_cat_name from db_category_cat_relation_temp where old_cat_id ="+old_cat_id

    cat_data = dict(dao.db.get_data(cat_sql)[0])

    goods_relation_data = {
        'goods_id': goods_id,
        'old_cat_id': cat_data['old_cat_id'],
        'old_cat_name': cat_data['old_cat_name'],
        'new_third_cat_id': cat_data['new_third_cat_id'],
        'new_third_cat_name': cat_data['new_third_cat_name'],
        'goods_status': '3'
    }

    dao.insert_without_exit("db_category_goods_relation_temp",goods_relation_data,goods_relation_data)
    print goods_data
    update_string.append("update db_goods set cat_id = "+str(cat_data['new_third_cat_id'])+" ,gmt_modified = @now_time where goods_id = "+goods_id)







file_object = open(r'/Users/zxg/Documents/work/淘气档口/新老分类替换/updateGoods.txt', 'w')

try:
    file_object.writelines("\n".join(update_string))
finally:
    file_object.close()
'''

'''# 删除不存在的goods 关系
goods_list = list()
goods_sql = "select goods_id from db_goods"
for goods_data in dao.db.get_data(goods_sql):
    goods_list.append(str(goods_data['goods_id']))

not_list = list()
goods_relation_sql =  "select goods_id from db_category_goods_relation_temp group by goods_id"
for goods_relation in dao.db.get_data(goods_relation_sql):
    goods_id = str(goods_relation['goods_id'])
    if goods_id not in goods_list:
        not_list.append(goods_id)

    if len(not_list) == 5000:

        delete_string = "delete from db_category_goods_relation_temp where goods_id in ("+",".join(not_list)+")"
        dao.db.update_data(delete_string)
        not_list = list()

if len(not_list) > 0:

    delete_string = "delete from db_category_goods_relation_temp where goods_id in ("+",".join(not_list)+")"
    dao.db.update_data(delete_string)
'''

'''
wb = Workbook()
ew = ExcelWriter(workbook=wb)

cat_sheet = wb.create_sheet(u"attr", 0)

cat_sheet.cell(row=1, column=1).value = u'old二级id'
cat_sheet.cell(row=1, column=2).value = u'new三级id'
cat_sheet.cell(row=1, column=3).value = u'老一级'
cat_sheet.cell(row=1, column=4).value = u'老二级'
cat_sheet.cell(row=1, column=5).value = u'新一级'
cat_sheet.cell(row=1, column=6).value = u'新二级'
cat_sheet.cell(row=1, column=7).value = u'新三级'

old_cat_dict = dict()
new_cat_dict = dict()
cat_dict = dict()
cate_sql = "select cat_id,cat_name,parent_id,cat_code,vehicle_code from db_category "
for cate_data in dao.db.get_data(cate_sql):
    cat_id = str(cate_data['cat_id'])

    cat_name = str(cate_data['cat_name'])
    parent_id = str(cate_data['parent_id'])

    cat_dict[cat_id] = cate_data
    if int(cat_id) < 3000:
        old_key = cat_name+parent_id
        old_cat_dict[cat_id] = cate_data
    if int(cat_id) > 2999:
        cat_code = str(cate_data['cat_code'])
        new_key = cat_code+"-"+parent_id
        new_cat_dict[new_key] = cate_data

cat_list = set()
true_cat_sql = "select old_cat_id from db_category_cat_relation_temp_true group by old_cat_id"
for true_data in dao.db.get_data(true_cat_sql):
    old_cat_id = str(true_data['old_cat_id'])
    cat_list.add(old_cat_id)
true_cat_sql = "select old_cat_id from db_category_cat_relation_temp_from_goods where cat_status = 4 group by old_cat_id"
for true_data in dao.db.get_data(true_cat_sql):
    old_cat_id = str(true_data['old_cat_id'])
    cat_list.add(old_cat_id)

row_num = 2
true_list =list()
true_cat_sql = "select cat_id from db_category where cat_id < 3000 and parent_id != 0"
for true_data in dao.db.get_data(true_cat_sql):
    old_cat_id = str(true_data['cat_id'])
    # new_third_cat_id = str(true_data['new_third_cat_id'])
    #

    if old_cat_id in cat_list:
        continue
    old_second_cat = cat_dict[old_cat_id]
    old_second_name = old_second_cat['cat_name']

    old_first_id = str(old_second_cat['parent_id'])
    old_first_name = cat_dict[old_first_id]['cat_name']
    #
    # new_third_data = cat_dict[new_third_cat_id]
    # new_third_name = new_third_data['cat_name']
    #
    # new_second_data = cat_dict[str(new_third_data['parent_id'])]
    # new_second_name = new_second_data['cat_name']
    #
    # new_first_data = cat_dict[str(new_second_data['parent_id'])]
    # new_first_name = new_first_data['cat_name']

    cat_sheet.cell(row=row_num, column=1).value = old_cat_id
    # cat_sheet.cell(row=row_num, column=2).value = new_third_cat_id
    cat_sheet.cell(row=row_num, column=2).value = ''
    cat_sheet.cell(row=row_num, column=3).value = old_first_name
    cat_sheet.cell(row=row_num, column=4).value = old_second_name
    # cat_sheet.cell(row=row_num, column=5).value = new_first_name
    # cat_sheet.cell(row=row_num, column=6).value = new_second_name
    # cat_sheet.cell(row=row_num, column=7).value = new_third_name

    row_num += 1


ew.save(filename=r'/Users/zxg/Documents/work/淘气档口/新老分类替换/不存在新老对应cat.xlsx')
'''
'''
old_cat_dict = dict()
new_cat_dict = dict()
cat_dict = dict()
cate_sql = "SELECT cat_id,cat_name,parent_id,cat_code,vehicle_code,cat_level FROM db_category "

for cate_data in dao.db.get_data(cate_sql):
    cat_id = str(cate_data['cat_id'])

    cat_name = str(cate_data['cat_name'])
    parent_id = str(cate_data['parent_id'])
    cat_level = str(cate_data['cat_level'])
    vehicle_code = str(cate_data['vehicle_code'])

    cat_dict[cat_id] = cat_name
    if int(cat_id) < 3000:
        old_key = cat_name + parent_id
        old_cat_dict[old_key] = cat_id
    if int(cat_id) > 2999:
        cat_code = str(cate_data['cat_code'])
        new_key = cat_code + "-" + parent_id
        if cat_level == '3':
            new_key += vehicle_code
        new_cat_dict[new_key] = cate_data

excle_file = r'/Users/zxg/Documents/work/PythonExcle/标准分类体系excle/何先静.xls'
# 单个excle处理
data = fileDao.open_excel(excle_file)
#
table = data.sheets()[0]

n_rows = table.nrows  # 行数
n_cols = table.ncols  # 列数

print ('行数：%s ,列数：%s' % (n_rows, n_cols))
# 第二行开始
for row_num in range(1, n_rows):
    row = table.row_values(row_num)

    old_second_id = str(row[0]).strip().replace(".0", "").replace("（", "(").replace("）", ")")
    old_second_cat_name = cat_dict[old_second_id]

    new_code = str(row[7]).strip().replace(".0", "").replace("（", "(").replace("）", ")")
    if new_code == '':
        continue

        # old_first_cat_name = str(row[0]).strip().replace(".0", "").replace("（", "(").replace("）", ")")
        # old_second_cat_name = str(row[1]).strip().replace(".0", "").replace("（", "(").replace("）", ")")


        # old_first_key = old_first_cat_name+"0"
        # old_first_id = old_cat_dict[old_first_key]
        # old_second_key = old_second_cat_name+old_first_id
        # old_second_id = old_cat_dict[old_second_key]

    vehicle = 'C'

    first_code = new_code[0:1]
    second_code = new_code[1:3]
    third_code = new_code[3:6]

    first_key = first_code + "-" + "0"
    first_id = str(new_cat_dict[first_key]['cat_id'])
    second_id = str(new_cat_dict[second_code + "-" + first_id]['cat_id'])

    third_key = third_code + "-" + second_id + vehicle

    if third_key not in new_cat_dict.keys():
        # print third_key
        third_key = third_code + "-" + second_id + "H"
        if third_key not in new_cat_dict.keys():
            print 'again'

        continue
    third_data = new_cat_dict[third_code + "-" + second_id + vehicle]

    third_id = third_data['cat_id']
    third_name = third_data['cat_name']

    insert_data = {
        'old_cat_id': old_second_id,
        'old_cat_name': old_second_cat_name,
        'new_third_cat_id': third_id,
        'new_third_cat_name': third_name,
        'vehicle_code': vehicle
    }
    # print insert_data
    dao.insert_without_exit("db_category_cat_relation_temp", insert_data, insert_data)
'''

'''# 筛选出多个对应关系的old——cat
have_cat_list = list()
more_cat_list = set()


# cat_relation_sql = "select old_cat_id,old_cat_name,new_third_cat_id,new_third_cat_name from db_category_cat_relation_temp where cat_status = 0 group by old_cat_id"
# for cat_relation_data in dao.db.get_data(cat_relation_sql):
#     old_cat_id = str(cat_relation_data['old_cat_id'])
#     have_cat_list.append(old_cat_id)


cat_relation_sql = "select old_cat_id,old_cat_name,new_third_cat_id,new_third_cat_name from db_category_cat_relation_temp where cat_status = 0 group by old_cat_id,new_third_cat_id"
for cat_relation_data in dao.db.get_data(cat_relation_sql):
    old_cat_id = str(cat_relation_data['old_cat_id'])
    if old_cat_id not in have_cat_list:
        have_cat_list.append(old_cat_id)
    else:
        print old_cat_id



# file_object = open(r'/Users/zxg/Documents/work/淘气档口/新老分类替换/more.txt', 'w')
#
# try:
#     file_object.writelines(",".join(more_cat_list))
# finally:
#     file_object.close()
'''

# 未整理对应关系的三级
'''# 将没处理的goods 均置为 5307

exist_id = '5307'
cate_dict = dict()
cate_sql = "SELECT cat_id,cat_name,parent_id,cat_code,vehicle_code,cat_level FROM db_category"
for cate_data in dao.db.get_data(cate_sql):
    cat_id = str(cate_data['cat_id'])
    cat_name = str(cate_data['cat_name'])
    cate_dict[cat_id] = cat_name

goods_relation_list = list()
goods_relation_sql = "select goods_id from db_category_goods_relation_temp "
for goods_relation_data in dao.db.get_data(goods_relation_sql):
    goods_id = str(goods_relation_data['goods_id'])
    goods_relation_list.append(goods_id)

batch_list = list()
goods_sql = "select goods_id,cat_id from db_goods "
for goods_data in dao.db.get_data(goods_sql):
    goods_id = str(goods_data['goods_id'])
    old_cat_id = str(goods_data['cat_id'])
    if goods_id not in goods_relation_list:
        goods_data = {
            'goods_id':goods_id,
            'old_cat_id':old_cat_id,
            'new_third_cat_id':exist_id,
            'old_cat_name': cate_dict[old_cat_id],
            'new_third_cat_name': cate_dict[exist_id]
        }
        batch_list.append(goods_data)
    if len(batch_list) > 5000:
        dao.insert_batch_temple("db_category_goods_relation_temp",batch_list)
        batch_list = list()

if len(batch_list) > 0:
    dao.insert_batch_temple("db_category_goods_relation_temp",batch_list)
'''

'''# 没有匹配的商品均置为 5307-回收站,先生成up数据
exist_id = '5307'

update_list = list()
update_list.append("select @now_time := now();")

third_dict = dict()
goods_relation_sql = "select goods_id,new_third_cat_id from db_category_goods_relation_temp ORDER BY new_third_cat_id"
for goods_relation_data in dao.db.get_data(goods_relation_sql):
    goods_id = str(goods_relation_data['goods_id'])
    new_third_cat_id = str(goods_relation_data['new_third_cat_id'])

    # if new_third_cat_id == exist_id:
    #     continue

    if new_third_cat_id in third_dict.keys():
        goods_set = set(third_dict[new_third_cat_id])
    else:
        goods_set = set()

    goods_set.add(goods_id)
    third_dict[new_third_cat_id] = goods_set

for new_third_cat_id in third_dict.keys():
    goods_set = set(third_dict[new_third_cat_id])

    string = "update db_goods set cat_id = " + new_third_cat_id + ",gmt_modified = @now_time where goods_id in (" + ",".join(goods_set) + ");"
    update_list.append(string)

file_object = open(r'/Users/zxg/Documents/work/淘气档口/新老分类替换/updateGoodsCate.txt', 'w')

# update_list.append("update db_goods set cat_id = "+exist_id+",gmt_modified = @now_time where cat_id < 3000;")
try:
    file_object.writelines("\n".join(update_list))
finally:
    file_object.close()
'''

'''# 生成cat name数据

cate_dict = dict()
cate_sql = "SELECT cat_id,cat_name,parent_id,cat_code,vehicle_code,cat_level FROM db_category "
for cate_data in dao.db.get_data(cate_sql):
    cat_id = str(cate_data['cat_id'])
    cat_name = str(cate_data['cat_name'])
    cate_dict[cat_id] = cat_name

goods_relation_sql = "SELECT old_cat_id,new_third_cat_id FROM db_category_goods_relation_temp group by old_cat_id,new_third_cat_id"
for goods_data in dao.db.get_data(goods_relation_sql):
    old_cat_id = str(goods_data['old_cat_id'])
    new_third_cat_id = str(goods_data['new_third_cat_id'])

    new_third_cat_name = cate_dict[new_third_cat_id]
    old_cat_name = cate_dict[old_cat_id]
    update_data = {
        'old_cat_name': old_cat_name,
        'new_third_cat_name': new_third_cat_name
    }
    exist_data = {
        'old_cat_id':old_cat_id,
        'new_third_cat_id':new_third_cat_id
    }

    dao.update_temple("db_category_goods_relation_temp", update_data, exist_data)
'''

'''
# 处理问题数据修正
# excle_file = r'/Users/zxg/Documents/work/PythonExcle/标准分类体系excle/问题数据修正.xlsx'
excle_file = r'/Users/zxg/Documents/work/PythonExcle/标准分类体系excle/电商商品转换-BOSHI.xlsx'
# 单个excle处理
data = fileDao.open_excel(excle_file)
#
table = data.sheets()[0]

n_rows = table.nrows  # 行数
n_cols = table.ncols  # 列数

print ('行数：%s ,列数：%s' % (n_rows, n_cols))
# 第二行开始
# 分类
cate_dict = dict()
cate_key_dict = dict()
cate_sql = "SELECT cat_id,cat_name,parent_id,cat_code,vehicle_code,cat_level FROM db_category_dian WHERE cat_id > 2999"
for cate_data in dao.db.get_data(cate_sql):
    cat_id = str(cate_data['cat_id'])
    cat_name = str(cate_data['cat_name'])
    cate_dict[cat_id] = cat_name

    key = str(cate_data['parent_id']) + str(cate_data['cat_code'])
    cate_key_dict[key] = cat_id

goods_dict = dict()
for row_num in range(1, n_rows):
    row = table.row_values(row_num)

    goods_id = str(row[0]).strip().replace(".0", "").replace("（", "(").replace("）", ")")
    # result = str(row[6]).strip().replace(".0", "").replace("（", "(").replace("）", ")")
    new_code = str(row[19]).strip().replace(".0", "").replace("（", "(").replace("）", ")")

    first_code = new_code[0:1]
    second_code = new_code[1:3]
    third_code = new_code[3:6]

    # if result != '错误':
    #     continue

    first_key = "0" + first_code
    first_id = cate_key_dict[first_key]

    second_key = first_id + second_code
    second_id = cate_key_dict[second_key]

    third_key = second_id + third_code
    third_id = cate_key_dict[third_key]

    update_data = {
        'new_third_cat_id': third_id,
        'new_third_cat_name': cate_dict[third_id]
    }

    dao.update_temple("db_category_goods_relation_temp",update_data,{'goods_id': goods_id})
'''

'''# start 处理新增的商品数据

def save_cat(cat_name, cat_kind, cat_level, cat_code, vehicle_code,parent_id):
    if cat_code == 'A':
        sort_order = 7
    elif cat_code == 'B':
        sort_order = 8
    elif cat_code == 'C':
        sort_order = 9
    elif cat_code == 'D':
        sort_order = 10
    elif cat_code == 'Z':
        sort_order = 11
    elif cat_code == 'XXX':
        sort_order = 999
    else:
        sort_order = int(cat_code)

    cat_data = {
        'cat_name': cat_name,
        'cat_kind': cat_kind,
        'cat_level': cat_level,
        'cat_code': cat_code,
        'vehicle_code': vehicle_code,
        'category_thumb': '',
        'category_img': '',
        'original_img': '',
        'style': '',
        'sort_order': sort_order,
        'parent_id':parent_id
    }

    return str(dao.insert_without_exit("db_category_dian", cat_data, cat_data, cat_id))


excle_file = r'/Users/zxg/Documents/work/PythonExcle/标准分类体系excle/配件标准模板v105版（碰撞+车类）.xlsx'

# 分类
cate_dict = dict()
cate_key_dict = dict()
cate_sql = "SELECT cat_id,cat_name,parent_id,cat_code,vehicle_code,cat_level FROM db_category_dian WHERE cat_id > 2999"
for cate_data in dao.db.get_data(cate_sql):
    cat_id = str(cate_data['cat_id'])
    cate_dict[cat_id] = cate_data

    key = str(cate_data['cat_name']) + str(cate_data['parent_id']) + str(cate_data['cat_code'])

    if str(cate_data['cat_level']) == '3':
        key += str(cate_data['vehicle_code'])

    cate_key_dict[key] = cat_id


# 单个excle处理
data = fileDao.open_excel(excle_file)
#
table = data.sheets()[0]

n_rows = table.nrows  # 行数
n_cols = table.ncols  # 列数

print ('行数：%s ,列数：%s' % (n_rows, n_cols))
# 第二行开始
cat_kind_dict = {'全车件': '1', '易损件': '0'}

goods_dict = dict()
for row_num in range(1, n_rows):
    row = table.row_values(row_num)

    first_name = str(row[0]).strip().replace(".0", "").replace("（", "(").replace("）", ")")
    second_name = str(row[2]).strip().replace(".0", "").replace("（", "(").replace("）", ")")
    third_name = str(row[4]).strip().replace(".0", "").replace("（", "(").replace("）", ")")
    sum_code = str(row[8]).strip().replace(".0", "").replace("（", "(").replace("）", ")")
    cat_kind_name = str(row[11]).strip().replace(".0", "").replace("（", "(").replace("）", ")")
    vehicle_code = str(row[13]).strip().replace(".0", "").replace("（", "(").replace("）", ")")

    first_code = sum_code[0:1]
    second_code = sum_code[1:3]
    third_code = sum_code[3:6]

    first_key = first_name + "0" + first_code
    if first_key not in cate_key_dict.keys():
        first_id = save_cat(first_name, '2', '1', first_code, vehicle_code,'0')
        cate_key_dict[first_key] = first_id
    else:
        first_id = cate_key_dict[first_key]

    try:
        second_key = second_name + first_id + second_code
    except:
        print 'hh'
    if second_key not in cate_key_dict.keys():
        second_id = save_cat(second_name, '2', '2', second_code, vehicle_code,first_id)
        cate_key_dict[second_key] = second_id
    else:
        second_id = cate_key_dict[second_key]

    if vehicle_code == 'CH':
        for vehi in ('C', 'H'):
            try:
                third_key = third_name + second_id + third_code + vehi
            except:
                print 'hh'
            if third_key not in cate_key_dict.keys():
                cat_kind = cat_kind_dict[cat_kind_name]
                third_id = save_cat(third_name, cat_kind, '3', third_code, vehi,second_id)
                cate_key_dict[third_key] = third_id
            else:
                third_id = cate_key_dict[third_key]
    else:
        third_key = third_name + second_id + third_code + vehicle_code
        if third_key not in cate_key_dict.keys():
            cat_kind = cat_kind_dict[cat_kind_name]
            third_id = save_cat(third_name, cat_kind, '3', third_code, vehicle_code,second_id)
            cate_key_dict[third_key] = third_id
        else:
            third_id = cate_key_dict[third_key]

'''  # end 处理新增的商品数据

'''
# 处理db_category_attribute 属性对应关系导出excle
wb = Workbook()
ew = ExcelWriter(workbook=wb)

attr_sheet = wb.create_sheet(u"attr", 0)

attr_sheet.cell(row=1, column=1).value = u'车辆种类码'
attr_sheet.cell(row=1, column=2).value = u'一级'
attr_sheet.cell(row=1, column=3).value = u'二级'
attr_sheet.cell(row=1, column=4).value = u'三级id'
attr_sheet.cell(row=1, column=5).value = u'三级'
attr_sheet.cell(row=1, column=6).value = u'属性id'
attr_sheet.cell(row=1, column=7).value = u'属性名'
attr_sheet.cell(row=1, column=8).value = u'状态'


# 分类
cate_dict = dict()
cate_sql = "SELECT cat_id,cat_name,parent_id,vehicle_code FROM db_category WHERE cat_id > 2999"
for cate_data in dao.db.get_data(cate_sql):
    cate_dict[str(cate_data['cat_id'])] = cate_data

# 属性
cat_attr_dict = dict()
cat_attr_sql = "SELECT dca.cat_id as cat_id,dca.attr_id as attr_id , dac.attr_name as attr_name FROM db_category_attribute dca,db_attribute_config dac WHERE dca.attr_id = dac.id AND dca.is_deleted = 'N'"
for cat_attr_data in dao.db.get_data(cat_attr_sql):
    cat_id = str(cat_attr_data['cat_id'])
    attr_id = str(cat_attr_data['attr_id'])
    attr_name = str(cat_attr_data['attr_name'])

    # attr_vo = {'attr_id':attr_id,"attr_name":attr_name}
    attr_vo = attr_id+"-"+attr_name
    if cat_id in cat_attr_dict.keys():
        attr_set = set(cat_attr_dict[cat_id])
    else:
        attr_set = set()
    attr_set.add(attr_vo)
    cat_attr_dict[cat_id] = attr_set


third_cat_attr_dict = dict()
# 出现过的三级id
all_have_new_cat_set = set()
all_have_old_cat_set = set()

cat_relation_sql = "SELECT old_cat_id,new_third_cat_id FROM db_category_cat_relation_temp "
cat_array = dao.db.get_data(cat_relation_sql)
for cat_data in cat_array:
    old_cat_id = str(cat_data['old_cat_id'])
    new_third_cat_id = str(cat_data['new_third_cat_id'])

    all_have_new_cat_set.add(new_third_cat_id)

    # 不存在分类
    if old_cat_id not in cat_attr_dict.keys():
        continue

    if new_third_cat_id == '4143':
        print 'hh'
    if new_third_cat_id in third_cat_attr_dict.keys():
        third_attr_set = set(third_cat_attr_dict[new_third_cat_id])
    else:
        third_attr_set = set()

    for attr_vo in cat_attr_dict[old_cat_id]:
        third_attr_set.add(attr_vo)

    third_cat_attr_dict[new_third_cat_id] = third_attr_set
    all_have_old_cat_set.add(old_cat_id)

# 写
insert_string = ''
insert_list = list()
insert_string += "select @now_time := now();\n"
insert_string += "update db_category_attribute set is_deleted = 'Y';\n"
insert_string += "insert into db_category_attribute(gmt_create,gmt_modified,cat_id,attr_id) VALUE \n"

row_num = 2
for new_third_cat_id in third_cat_attr_dict.keys():
    if new_third_cat_id == '4143':
        print 'hh'
    third_data = cate_dict[new_third_cat_id]
    second_data = cate_dict[str(third_data['parent_id'])]
    first_data = cate_dict[str(second_data['parent_id'])]

    third_attr_set = set(third_cat_attr_dict[new_third_cat_id])
    if len(third_attr_set) == 0:
        # 无分类属性
        continue
    else:
        for attr_vo in third_attr_set:
            attr_array = attr_vo.split("-")

            insert_list.append("(@now_time,@now_time,'"+new_third_cat_id+"' , '"+str(attr_array[0])+"' )")
#
for old_cat_id in cat_attr_dict.keys():
    if old_cat_id not in all_have_old_cat_set:
        print old_cat_id

insert_string += ",".join(insert_list)
insert_string += ";"
file_object = open(r'/Users/zxg/Documents/work/淘气档口/新老分类替换/insetAttr.txt', 'w')

try:
    file_object.writelines(insert_string)
finally:
    file_object.close()
'''

# # 不存在的三级分类
# not_cat_sql = "select cat_id,cat_name,cat_code,vehicle_code from db_category where cat_id > 2999 and cat_level = 3"
# for not_data in dao.db.get_data(not_cat_sql):
#     cat_id = str(not_data['cat_id'])
#     if cat_id not in all_have_new_cat_set:
#         third_data = cate_dict[cat_id]
#         second_data = cate_dict[str(third_data['parent_id'])]
#         first_data = cate_dict[str(second_data['parent_id'])]
#         attr_sheet.cell(row=row_num, column=1).value = third_data['vehicle_code']
#         attr_sheet.cell(row=row_num, column=2).value = first_data['cat_name']
#         attr_sheet.cell(row=row_num, column=3).value = second_data['cat_name']
#         attr_sheet.cell(row=row_num, column=4).value = cat_id
#         attr_sheet.cell(row=row_num, column=5).value = third_data['cat_name']
#         attr_sheet.cell(row=row_num, column=6).value = '0'
#         attr_sheet.cell(row=row_num, column=7).value = ''
#         attr_sheet.cell(row=row_num, column=8).value = u'无老分类'
#         row_num += 1
#
#
# ew.save(filename=r'/Users/zxg/Documents/work/淘气档口/新老分类替换/ttr.xlsx')


'''
# 所有出现的三级分类
all_have_new_cat_set = set()

wb = Workbook()
ew = ExcelWriter(workbook=wb)

equal_sheet = wb.create_sheet(u"1-1", 0)
second_sheet = wb.create_sheet(u"1-n", 1)
third_sheet = wb.create_sheet(u"n-1", 2)
not_show_sheet = wb.create_sheet(u"not", 3)
mapping_sheet = wb.create_sheet(u"not in mapping", 4)


mapping_sheet.cell(row=1, column=1).value = u'third_cat_id'
mapping_sheet.cell(row=1, column=2).value = u'third_cat_name'

not_show_sheet.cell(row=1, column=1).value = u'third_cat_id'
not_show_sheet.cell(row=1, column=2).value = u'third_cat_name'
not_show_sheet.cell(row=1, column=3).value = u'third_cat_code'
not_show_sheet.cell(row=1, column=4).value = u'vehicle_code'


equal_sheet.cell(row=1, column=1).value = u'old_cat_id'
equal_sheet.cell(row=1, column=2).value = u'old_name'
equal_sheet.cell(row=1, column=3).value = u'new_third_cat_id'
equal_sheet.cell(row=1, column=4).value = u'new_name'

second_sheet.cell(row=1, column=1).value = u'old_cat_id'
second_sheet.cell(row=1, column=2).value = u'old_name'
second_sheet.cell(row=1, column=3).value = u'new_third_cat_id'
second_sheet.cell(row=1, column=4).value = u'new_name'

third_sheet.cell(row=1, column=1).value = u'old_cat_id'
third_sheet.cell(row=1, column=2).value = u'old_name'
third_sheet.cell(row=1, column=3).value = u'new_third_cat_id'
third_sheet.cell(row=1, column=4).value = u'new_name'

# 为excel额外数据服务
cate_dict = dict()
cate_sql = "select cat_id,cat_name from db_category"
for cate_data in dao.db.get_data(cate_sql):
    cate_dict[str(cate_data['cat_id'])] = str(cate_data['cat_name'])

# 检验mapping表是否是全部在售商品cat
mapping_list = list()
mapping_sql = "select cat_id from db_category_mapping where cat_mapping_level = 2"
for mapping_data in dao.db.get_data(mapping_sql):
    mapping_list.append(str(mapping_data['cat_id']))
not_in_mapping_list = list()


old_to_new_dict = dict()
new_to_old_dict = dict()
cat_relation_sql = "select old_cat_id,new_third_cat_id from db_category_cat_relation_temp where cat_status = 3"
cat_array = dao.db.get_data(cat_relation_sql)
for cat_data in cat_array:
    old_cat_id = str(cat_data['old_cat_id'])
    new_third_cat_id = str(cat_data['new_third_cat_id'])

    all_have_new_cat_set.add(new_third_cat_id)

    if new_third_cat_id not in mapping_list:
        not_in_mapping_list.append(new_third_cat_id)

    # 存 old：list（new）
    if old_cat_id in old_to_new_dict.keys():
        new_list = list(old_to_new_dict[old_cat_id])
    else:
        new_list = list()
    new_list.append(new_third_cat_id)
    old_to_new_dict[old_cat_id] = new_list

    # 存 new：list（old）
    if new_third_cat_id in new_to_old_dict.keys():
        old_list = list(new_to_old_dict[new_third_cat_id])
    else:
        old_list = list()
    old_list.append(old_cat_id)
    new_to_old_dict[new_third_cat_id] = old_list

# 提取1：1
old_new_one_one_key_list = list()
old_new_one_one_key_dict = dict()
# 找出1：n的数据
second_num = 2
for old_cat_id in old_to_new_dict.keys():
    new_list = list(old_to_new_dict[old_cat_id])
    if len(new_list) == 1:
        key = old_cat_id+new_list[0]
        old_new_one_one_key_list.append(key)
    else:
        for new_third_id in new_list:
            second_sheet.cell(row=second_num, column=1).value = old_cat_id
            second_sheet.cell(row=second_num, column=2).value = cate_dict[old_cat_id]
            second_sheet.cell(row=second_num, column=3).value = new_third_id
            second_sheet.cell(row=second_num, column=4).value = cate_dict[new_third_id]
            second_num +=1

# 找出n:1
third_num = 2
for new_third_cat_id in new_to_old_dict.keys():
    old_list = list(new_to_old_dict[new_third_cat_id])
    if len(old_list) == 1:
        key = old_list[0]+new_third_cat_id
        if key in old_new_one_one_key_list:
            old_new_one_one_key_dict[old_list[0]] = new_third_cat_id
    else:
        for old_cat_id in old_list:
            third_sheet.cell(row=third_num, column=1).value = old_cat_id
            third_sheet.cell(row=third_num, column=2).value = cate_dict[old_cat_id]
            third_sheet.cell(row=third_num, column=3).value = new_third_cat_id
            third_sheet.cell(row=third_num, column=4).value = cate_dict[new_third_cat_id]
            third_num += 1

# 1:1
first_num = 2
for old_cat_id in old_new_one_one_key_dict.keys():
    new_third_cat_id = old_new_one_one_key_dict[old_cat_id]
    equal_sheet.cell(row=first_num, column=1).value = old_cat_id
    equal_sheet.cell(row=first_num, column=2).value = cate_dict[old_cat_id]
    equal_sheet.cell(row=first_num, column=3).value = new_third_cat_id
    equal_sheet.cell(row=first_num, column=4).value = cate_dict[new_third_cat_id]
    first_num += 1



# 找出不存在在对应关系的新分类
not_num = 2
not_cat_sql = "select cat_id,cat_name,cat_code,vehicle_code from db_category where cat_id > 2999 and cat_level = 3"
for not_data in dao.db.get_data(not_cat_sql):
    cat_id = str(not_data['cat_id'])
    if cat_id not in all_have_new_cat_set:
        not_show_sheet.cell(row=not_num, column=1).value = cat_id
        not_show_sheet.cell(row=not_num, column=2).value = str(not_data['cat_name'])
        not_show_sheet.cell(row=not_num, column=3).value = str(not_data['cat_code'])
        not_show_sheet.cell(row=not_num, column=4).value = str(not_data['vehicle_code'])
        not_num += 1

# mapping_data
mapping_num = 2
for third_cat_id in not_in_mapping_list:
    mapping_sheet.cell(row=mapping_num, column=1).value = third_cat_id
    mapping_sheet.cell(row=mapping_num, column=2).value = str(cate_dict[third_cat_id])
    mapping_num += 1


ew.save(filename=r'/Users/zxg/Documents/work/淘气档口/新老分类替换/result.xlsx')
'''

'''# 更新relation goods 状态
not_up_list = list()
up_list = list()
goods_sql = "select cat_id,goods_id,goods_name from db_goods where seller_id = 1 and is_delete = 0"
goods_array = dao.db.get_data(goods_sql)
for goods_data in goods_array:
    goods_id = str(goods_data['goods_id'])
    goods_name = str(goods_data['goods_name'])

    # 判断有无库存
    stock_sql = "select goods_number,stock_type,immediate_stock from db_goods_stock where goods_id = '" + str(goods_id) + "' and is_deleted = 'N' and seller_id = 1"
    stock_array = dao.db.get_data(stock_sql)
    if len(stock_array) == 0:
        not_up_list.append(goods_id)
        continue
    number = 0

    #
    for stock_result in stock_array:
        stock_type = int(stock_result['stock_type'])
        if stock_type == 0:
            number += int(stock_result['goods_number'])
        else:
            number += int(stock_result['immediate_stock'])

    if str(number) == '0':
        not_up_list.append(goods_id)
        continue

    # 判断是否在售
    warehouse_select_sql = "select seller_id,seller_nick,is_on_sale from db_goods_warehouse where goods_id = '" + str(goods_id) + "' and seller_id = 1 and is_on_sale = 1"
    warehouse_array = dao.db.get_data(warehouse_select_sql)
    if len(warehouse_array) == 0:
        not_up_list.append(goods_id)
        continue
    up_list.append(goods_id)


up_sql = "update db_category_goods_relation_temp set goods_status = 3 where goods_id in ("+",".join(up_list)+")"
not_up_sql = "update db_category_goods_relation_temp set goods_status = 2 where goods_id in ("+",".join(not_up_list)+")"

dao.db.update_data(up_sql)
dao.db.update_data(not_up_sql)
'''

'''
# 查找出cat 数据 插入
exist_id = '5307'
insert_list = list()
exist_list = list()
for goods_status in ('3', '2', '1'):
    relation_sql = "select old_cat_id,old_cat_name,new_third_cat_id,new_third_cat_name,goods_status from db_category_goods_relation_temp where goods_status = "+goods_status+" group by old_cat_id,new_third_cat_id"
    relation_array = dao.db.get_data(relation_sql)
    for relation_data in relation_array:
        old_cat_id = str(relation_data['old_cat_id'])
        old_cat_name = str(relation_data['old_cat_name'])
        new_third_cat_id = str(relation_data['new_third_cat_id'])
        new_third_cat_name = str(relation_data['new_third_cat_name'])

        # 不存回收站数据
        if new_third_cat_id == exist_id:
            continue

        key = old_cat_id+"-"+new_third_cat_id
        if key in exist_list:
            continue
        else:
            exist_list.append(key)

        cat_data = {
            'old_cat_id':old_cat_id,
            'old_cat_name':old_cat_name,
            'new_third_cat_id':new_third_cat_id,
            'new_third_cat_name':new_third_cat_name,
            'cat_status':goods_status,
        }
        insert_list.append(cat_data)


dao.insert_batch_temple("db_category_cat_relation_temp", insert_list)
'''

'''
all_cate_dict = dict()
all_cate_sql = "select cat_id,cat_name from db_category "
all_cate_array = dao.db.get_data(all_cate_sql)
for all_cate_data in all_cate_array:
    all_cate_dict[str(all_cate_data['cat_id'])] = str(all_cate_data['cat_name'])
# 电商1：n的

goods_dict = dict()
cate_sql = "SELECT old_cat_id,new_third_cat_id,goods_id FROM db_category_relation "
cate_relation_array = dao.db.get_data(cate_sql)
for cate_relation_data in cate_relation_array:

    goods_id = str(cate_relation_data['goods_id'])
    goods_dict[goods_id] = cate_relation_data

goods_name_dict = dict()
old_cat_dict = dict()
old_new_dict = dict()
goods_sql = "select cat_id,goods_id,goods_name from db_goods where seller_id = 1 and is_delete = 0"
goods_array = dao.db.get_data(goods_sql)
for goods_data in goods_array:
    goods_id = str(goods_data['goods_id'])
    goods_name = str(goods_data['goods_name'])

    # 判断有无库存
    stock_sql = "select goods_number from db_goods_stock where goods_id = '" + str(goods_id) + "' and is_deleted = 'N' and seller_id = 1"
    stock_array = dao.db.get_data(stock_sql)
    if len(stock_array) == 0:
        continue
    number = 0
    for stock_result in stock_array:
        number += int(stock_result['goods_number'])

    if str(number) == '0':
        continue

    # 判断是否在售
    warehouse_select_sql = "select seller_id,seller_nick,is_on_sale from db_goods_warehouse where goods_id = '" + str(goods_id) + "' and seller_id = 1 and is_on_sale = 1"
    warehouse_array = dao.db.get_data(warehouse_select_sql)
    if len(warehouse_array) == 0:
        continue

    goods_name_dict[goods_id] = goods_name
    cate_relation_data = goods_dict[goods_id]
    old_cat_id = str(cate_relation_data['old_cat_id'])
    new_third_cat_id = str(cate_relation_data['new_third_cat_id'])
    if old_cat_id in old_cat_dict.keys():
        cate_relation_set = set(old_cat_dict[old_cat_id])
    else:
        cate_relation_set = set()

    cate_relation_set.add(new_third_cat_id)
    old_cat_dict[old_cat_id] = cate_relation_set

    # 此时老新id对应的
    key = old_cat_id+"_"+new_third_cat_id
    if key in old_new_dict.keys():
        goods_id_set = set(old_new_dict[key])
    else:
        goods_id_set = set()

    goods_id_set.add(goods_id)
    old_new_dict[key] = goods_id_set

wb = Workbook()
ew = ExcelWriter(workbook=wb)

car_sheet = wb.create_sheet(u"重复", 0)
# 第一行
car_sheet.cell(row=1, column=1).value = u'goods_id'
car_sheet.cell(row=1, column=2).value = u'goods_name'
car_sheet.cell(row=1, column=3).value = u'旧分类id'
car_sheet.cell(row=1, column=4).value = u'旧分类name'
car_sheet.cell(row=1, column=5).value = u'新分类id'
car_sheet.cell(row=1, column=6).value = u'新分类name'

write_list = list()
row_insert_num = 2
for old_cat_id in old_cat_dict.keys():
    cate_relation_set = set(old_cat_dict[old_cat_id])
    # if len(cate_relation_set) == 1:
    #     continue

    cate_relation_set = set(old_cat_dict[old_cat_id])
    for new_third_cat_id in cate_relation_set:


        key = old_cat_id+"_"+new_third_cat_id
        goods_id_set = set(old_new_dict[key])
        car_sheet.cell(row=row_insert_num, column=5).value = ",".join(goods_id_set)

        for goods_id in goods_id_set:
            car_sheet.cell(row=row_insert_num, column=1).value = goods_id
            car_sheet.cell(row=row_insert_num, column=2).value = goods_name_dict[goods_id]
            car_sheet.cell(row=row_insert_num, column=3).value = old_cat_id
            car_sheet.cell(row=row_insert_num, column=4).value = all_cate_dict[old_cat_id]
            car_sheet.cell(row=row_insert_num, column=5).value = new_third_cat_id
            car_sheet.cell(row=row_insert_num, column=6).value = all_cate_dict[new_third_cat_id]

            row_insert_num += 1

ew.save(filename=r'/Users/zxg/Documents/work/淘气档口/新老分类替换/oneToMore.xlsx')
# write_update_file = r'/Users/zxg/Documents/work/淘气档口/新老分类替换/oneToMore1.txt'
# file_object = open(write_update_file, 'w')
#
# try:
#     file_object.writelines("\n\n\n".join(write_list))
# finally:
#     file_object.close()
'''

'''
# 分类加序号
cat_sql = "select * from db_category_dian where cat_id > 2999"
cate_array = dao.db.get_data(cat_sql)
for cate_data in cate_array:
    cat_code = str(cate_data['cat_code'])

    if cat_code == 'A':
        sort_order = 7
    elif cat_code == 'B':
        sort_order = 8
    elif cat_code == 'C':
        sort_order = 9
    elif cat_code == 'XXX':
        sort_order = 999
    else:
        sort_order = int(cat_code)

    cate_data['sort_order'] = sort_order
    cate_data['creator'] = 0
    cate_data['modifier'] = 0
    dao.insert_temple("db_category_new",cate_data)
'''


# 博士给到的在售商品 生成goods：新三级分类-excle
# excle_file = r'/Users/zxg/Documents/work/PythonExcle/标准分类体系excle/电商商品转化-20151203.xlsx'
# # 单个excle处理
# data = fileDao.open_excel(excle_file)
# #
# table = data.sheets()[0]
#
# n_rows = table.nrows  # 行数
# n_cols = table.ncols  # 列数
#
# print ('行数：%s ,列数：%s' % (n_rows, n_cols))
# # 第二行开始
#
# goods_dict = dict()
# for row_num in range(1, n_rows):
#     row = table.row_values(row_num)
#
#     goods_id = str(row[0]).strip().replace(".0", "")
#     vehicle_name = str(row[6]).strip().replace(".0", "")
#     first_name = str(row[11]).strip()
#     second_name = str(row[12]).strip()
#     third_name = str(row[13]).strip()
#     goods_data = {
#         'vehicle_name': vehicle_name,
#         'first_name': first_name,
#         'second_name': second_name,
#         'third_name': third_name
#     }
#     goods_dict[goods_id] = goods_data
#
# wb = Workbook()
# ew = ExcelWriter(workbook=wb)
#
# car_sheet = wb.create_sheet(u"分类", 0)
# # 第一行
# car_sheet.cell(row=1, column=1).value = u'商品id'
# car_sheet.cell(row=1, column=2).value = u'一级分类'
# car_sheet.cell(row=1, column=3).value = u'二级分类'
# car_sheet.cell(row=1, column=4).value = u'三级分类'
# car_sheet.cell(row=1, column=5).value = u'种类'
#
# excle_file1 = r'/Users/zxg/Documents/work/PythonExcle/标准分类体系excle/在售商品.xlsx'
# # 单个excle处理
# data1 = fileDao.open_excel(excle_file1)
# #
# table1 = data1.sheets()[0]
#
# n_rows = table1.nrows  # 行数
# n_cols = table1.ncols  # 列数
#
# print ('行数：%s ,列数：%s' % (n_rows, n_cols))
# # 第二行开始
# row_insert_num = 2
# for row_num in range(1, n_rows):
#     row = table1.row_values(row_num)
#     goods_id = str(row[0]).strip().replace(".0", "")
#     goods_data = goods_dict[goods_id]
#
#     car_sheet.cell(row=row_insert_num, column=1).value = str(goods_id)
#     car_sheet.cell(row=row_insert_num, column=2).value = str(goods_data['first_name'])
#     car_sheet.cell(row=row_insert_num, column=3).value = str(goods_data['second_name'])
#     car_sheet.cell(row=row_insert_num, column=4).value = str(goods_data['third_name'])
#     car_sheet.cell(row=row_insert_num, column=5).value = str(goods_data['vehicle_name'])
#     row_insert_num += 1
#
# ew.save(filename=r'/Users/zxg/Documents/work/淘气档口/新老分类替换/show.xlsx')



'''
# 2015-12-01 新copy分类给永斌
cat_id_dict = dict()
show_id_list = list()
cate_sql = "SELECT cat_id,cat_name,parent_id,cat_level,cat_code,vehicle_code FROM db_category WHERE is_deleted = 'N' AND cat_id > 1262"
cate_array = dao.db.get_data(cate_sql)
for cate_data in cate_array:
    cat_id = cate_data['cat_id']
    cat_level = cate_data['cat_level']
    cat_id_dict[cat_id] = cate_data
    if int(cat_level) == 3:
        show_id_list.append(cat_id)

wb = Workbook()
ew = ExcelWriter(workbook=wb)

car_sheet = wb.create_sheet(u"分类", 0)
# 第一行
car_sheet.cell(row=1, column=1).value = u'一级分类'
car_sheet.cell(row=1, column=2).value = u'code'
car_sheet.cell(row=1, column=3).value = u'二级分类'
car_sheet.cell(row=1, column=4).value = u'code'
car_sheet.cell(row=1, column=5).value = u'三级分类'
car_sheet.cell(row=1, column=6).value = u'code'
car_sheet.cell(row=1, column=7).value = u'种类码'
car_sheet.cell(row=1, column=8).value = u'线上goods'

cat_id_set = set()
cat_goods_dict = dict()
goods_sql = "select cat_id,goods_id from db_goods where seller_id = 1 and is_delete = 0 "
goods_array = dao.db.get_data(goods_sql)
for goods_data in goods_array:
    goods_id = int(goods_data['goods_id'])
    cat_id = int(goods_data['cat_id'])
    if cat_id < 3000:
        continue

    # 判断有无库存
    stock_sql = "select goods_number from db_goods_stock where goods_id = '" + str(goods_id) + "' and is_deleted = 'N' and seller_id = 1"
    stock_array = dao.db.get_data(stock_sql)
    if len(stock_array) == 0:
        continue
    number = 0
    for stock_result in stock_array:
        number += int(stock_result['goods_number'])

    if str(number) == '0':
        continue

    # 判断是否在售
    warehouse_select_sql = "select seller_id,seller_nick,is_on_sale from db_goods_warehouse where goods_id = '" + str(goods_id) + "' and seller_id = 1 and is_on_sale = 1"
    warehouse_array = dao.db.get_data(warehouse_select_sql)
    if len(warehouse_array) == 0:
        continue

    cat_id_set.add(cat_id)
    if cat_id in cat_goods_dict.keys():
        goods_list = cat_goods_dict[cat_id]
    else:
        goods_list = list()
    goods_list.append(str(goods_id))
    cat_goods_dict[cat_id] = goods_list

row_num = 2
for cat_id in cat_id_set:
    third_data = cat_id_dict[cat_id]
    goods_list = cat_goods_dict[cat_id]
    second_data = cat_id_dict[third_data['parent_id']]
    first_data = cat_id_dict[second_data['parent_id']]

    car_sheet.cell(row=row_num, column=1).value = str(first_data['cat_name'])
    car_sheet.cell(row=row_num, column=2).value = str(first_data['cat_code'])
    car_sheet.cell(row=row_num, column=3).value = str(second_data['cat_name'])
    car_sheet.cell(row=row_num, column=4).value = str(second_data['cat_code'])
    car_sheet.cell(row=row_num, column=5).value = str(third_data['cat_name'])
    car_sheet.cell(row=row_num, column=6).value = str(third_data['cat_code'])
    car_sheet.cell(row=row_num, column=7).value = str(third_data['vehicle_code'])
    car_sheet.cell(row=row_num, column=8).value = str(",".join(goods_list))

    # string = str(first_data['cat_name']) + " " + str(first_data['cat_code']) + " "
    # string += str(second_data['cat_name']) + " " + str(second_data['cat_code']) + " "
    # string += str(third_data['cat_name']) + " " + str(third_data['cat_code']) + " " + str(third_data['vehicle_code'])
    row_num += 1


# ew.save(filename=r'/Users/zxg/Desktop/test/yongbing.xlsx')  # # 分类三级是否准确
ew.save(filename=r'/Users/zxg/Documents/work/淘气档口/新老分类替换/show.xlsx')  # # 分类三级是否准确
'''
# file_object = open('/Users/zxg/Desktop/test/yongbing.txt', 'w')
#
# try:
#     file_object.writelines("\n".join(result_list))
# finally:
#     file_object.close()




# 其余行
# for cat_id in self.second_cat_list:
#
#     car_sheet.cell(row=row_num, column=1).value = self.cate_id_name[self.cate_id_parent[cat_id]]
#     car_sheet.cell(row=row_num, column=2).value = self.cate_id_name[cat_id]

# cate_set = list()
# cate_sql = "SELECT cat_name, cat_level FROM db_category_data WHERE is_deleted = 'N' AND cat_level = 3 and vehicle_code = 'CH'"
# cate_array = dao.db.get_data(cate_sql)
# for cate_data in cate_array:
#     cate_exist_data = {
#         'cat_name': str(cate_data['cat_name']),
#         'cat_level': str(cate_data['cat_level'])
#     }
#     cate_set.append(cate_exist_data)
#
# # 单个excle处理
# excle_file = r'/Users/zxg/Documents/work/PythonExcle/标准分类体系excle/配件标准模板v1.03版（碰撞+车类）.xlsx'
# data = fileDao.open_excel(excle_file)
# table = data.sheets()[0]
# n_rows = table.nrows  # 行数
# n_cols = table.ncols  # 列数
#
# print ('行数：%s ,列数：%s' % (n_rows, n_cols))
# # 第二行开始
# for row_num in range(1, n_rows):
#     row = table.row_values(row_num)
#
#     excle_vehicle_code = str(row[1].strip())
#     if excle_vehicle_code != 'CH':
#         continue
#     third_name = row[6].strip().replace("（", "(").replace("）", ")")
#     cate_exist_data = {
#         'cat_name': str(third_name),
#         'cat_level': '3'
#     }
#     if cate_exist_data not in cate_set:
#         print cate_exist_data

#
# result = a.compare(b)
# print type(result)
# if result == 0:
#     print '1'
# if result != 0:
#     print '2'
# print

'''
# 丹妮数据对接需求
excle_name = r'/Users/zxg/Desktop/danni/云修客户汇总表-20151130.xlsx'
# 单个excle处理
data = fileDao.open_excel(excle_name)
table = data.sheets()[1]
n_rows = table.nrows  # 行数
n_cols = table.ncols  # 列数

print ('行数：%s ,列数：%s' % (n_rows, n_cols))
# 第二行开始
phone_dict = dict()
for row_num in range(3, n_rows):
    row = table.row_values(row_num)

    phone_index = 11
    phone_num = str(row[phone_index]).strip().replace(".0", "")
    if phone_num == '':
        continue
    # if phone_num is None:

    brand_money = str(row[22]).strip()
    decorate_money = str(row[23]).strip()
    money = str(row[24]).strip()

    phone_dict[phone_num] = {
        'brand_money': brand_money,
        'decorate_money': decorate_money,
        'money': money
    }
    # print '%s-%s-%s-%s'(phone_num, brand_money, decorate_money, money)

people_dict = dict()
phone_set = set()
saint_customer_join_audit_sql = "SELECT contacts_mobile,id FROM saint_customer_join_audit WHERE is_deleted='N'"
audit_array = dao.db.get_data(saint_customer_join_audit_sql)
for audit_data in audit_array:
    id = str(audit_data['id'])
    contacts_mobile = str(audit_data['contacts_mobile'])
    people_dict[id] = contacts_mobile
    phone_set.add(contacts_mobile)

money_dict = dict()
money_sql = "SELECT customer_join_audit_id,pay_money,pay_type,is_confirm_account FROM saint_merchants_pay_money WHERE is_confirm_account = 1 AND is_deleted = 'N'"
money_array = dao.db.get_data(money_sql)
for money_data in money_array:
    customer_join_audit_id = str(money_data['customer_join_audit_id'])
    if customer_join_audit_id not in people_dict.keys():
        print 'error:此id不存在在用户表中：%s' % customer_join_audit_id
        continue
    phone = people_dict[customer_join_audit_id]
    money_has_data = {
        'pay_money': money_data['pay_money'],
        'pay_type': int(money_data['pay_type'])
    }
    if phone not in money_dict.keys():
        money_list = list()
    else:
        money_list = money_dict[phone]
    money_list.append(money_has_data)

    money_dict[phone] = money_list

# 1.手机号码不存在在数据库中
for phone in phone_dict.keys():
    if phone == '':
        continue
    if phone not in phone_set:
        print '1.手机号码不存在在saint_customer_join_audit 表中：%s' % phone
        continue

    excle_data = phone_dict[phone]
     # 品牌共建基金
    brand_money = excle_data['brand_money']
    if brand_money == '':
        brand_money = '0'
    excle_brand_money = Decimal(brand_money)

    # 装修保证金
    decorate_money = excle_data['decorate_money']
    if decorate_money == '':
        decorate_money = '0'
    excle_decorate_money = Decimal(decorate_money)

    if decorate_money == '0' and brand_money == '0':
        continue
    if phone not in money_dict.keys():
        print '2.手机号码不存在在 saint_merchants_pay_money 表中,无打款记录,而表中有记录：%s' % phone
        continue

    sql_money_list = money_dict[phone]


    has_brand_money = False
    has_decorate_money = False

    for sql_money_data in sql_money_list:
        pay_type = sql_money_data['pay_type']
        pay_money = sql_money_data['pay_money']

        if pay_type == 2:
            has_brand_money = True
            # 品牌共建金
            result = excle_brand_money.compare(pay_money)
            if result != 0 :
                print '4.手机号码品牌共建金不一致（excle：表）：%s-(%s:%s)' % (phone, str(excle_brand_money), str(pay_money))
        if pay_type == 3:
            has_decorate_money = True
            # 装修保证金
            result = excle_decorate_money.compare(pay_money)
            if result != 0:
                print '5.手机号码装修保证金不一致（excle：表）：%s-(%s:%s)' % (phone, str(excle_decorate_money), str(pay_money))

    if not has_brand_money and brand_money != '0':
        print '6.数据库手机号码无品牌共建金：%s-(%s)' % (phone, str(excle_brand_money))
    if not has_decorate_money and decorate_money != '0':
        print '7.数据库手机号码无装修保证金：%s-(%s)' % (phone, str(excle_decorate_money))

'''
# 临时处理产品库
# car_dict = dict()
# car_category = "select * from db_car_category where level = 6"
# car_array = dao.db.get_data(car_category)
# for car_data in car_array:
#     car_id = int(car_data['id'])
#     car_dict[car_id] = car_data
#
# li_car_sql = "select li_id from db_monkey_offer_li_car GROUP BY li_id"
# li_car_array = dao.db.get_data(li_car_sql)
# for li_car_data in li_car_array:
#     li_id = str(li_car_data['li_id'])
#     # liyang find online
#     liyang_sql = "select car_models_id from db_car_all where l_id = '"+li_id+"'"
#     car_models_id = int(dao.db.get_data(liyang_sql)[0]['car_models_id'])
#
#     car_data = car_dict[car_models_id]
#     li_car_update_data = {
#         'online_car_id':str(car_models_id),
#         'online_pid': str(car_data['pid']),
#         'online_brand': str(car_data['brand']),
#         'online_company': str(car_data['company']),
#         'online_series': str(car_data['series']),
#         'online_power': str(car_data['power']),
#         'status': '1'
#     }
#     li_car_exit_data ={
#         'li_id':li_id
#     }
#     dao.update_temple("db_monkey_offer_li_car",li_car_update_data,li_car_exit_data)

# print li_car_update_data


'''# 三级分类临时新增
def save_cat(cat_name, cat_code, parent_id, cat_kind, vehicle_code):

    sort_order = int(cat_code)

    cat_data = {
        'cat_name': cat_name,
        'cat_kind': cat_kind,
        'cat_level': '3',
        'cat_code': cat_code,
        'vehicle_code': vehicle_code,
        'category_thumb': '',
        'category_img': '',
        'original_img': '',
        'style': '',
        'sort_order': sort_order,
        'parent_id': parent_id
    }
    dao.insert_without_exit("db_category",cat_data,cat_data,'cat_id')

save_cat("发电机惰轮","117","4409",0,'C')
save_cat("发电机惰轮","117","4409",0,'H')
save_cat("车速传感器","310","3672",1,'C')
save_cat("车速传感器","310","3672",1,'H')
'''

'''# insert 线上的三级分类导出sql

file_object = open('/Users/zxg/Documents/work/淘气档口/新老分类替换/insertCategory.txt', 'w')
finally_list = list()
# finally_list.append("select @now_time := now();")
# finally_list.append("insert into db_category value ")

# 生成电商需要的sql
new_data_sql = "SELECT * FROM db_category where cat_id > 2999"
new_data_array = dao.db.get_data(new_data_sql)
for new_data in new_data_array:
    new_data = dict(new_data)
    new_data.pop("modifier")
    new_data.pop("creator")
    new_data.pop('gmt_create')
    new_data.pop('gmt_modified')
    insert_sql = dao.insert_temple("db_category", new_data)
    finally_list.append(insert_sql+";")

finally_list.append("update db_category set is_deleted = 'Y',gmt_modified=NOW() where cat_id  <  3000;")
try:
    file_object.writelines("\n".join(finally_list))
finally:
    file_object.close()
'''

# 线上三级分类 id 改为 3000
# old_new_id_dict = dict()
# all_id_list = list()
# new_data_sql = "SELECT * FROM db_category_dian WHERE cat_id > 1262 "
# new_data_array = dao.db.get_data(new_data_sql)
#
# start_index = 3000
# first = True
# for new_data in new_data_array:
#     new_data = dict(new_data)
#     new_data.pop("modifier")
#     new_data.pop("creator")
#     new_data.pop('gmt_create')
#     new_data.pop('gmt_modified')
#     cat_id = int(new_data['cat_id'])
#     parent_id = int(new_data['parent_id'])
#     cat_level = int(new_data['cat_level'])
#
#     old_cat_id = new_data['cat_id']
#     new_data.pop("cat_id")
#     if cat_level == 1:
#         new_data['parent_id'] = 0
#     else:
#         new_data['parent_id'] = old_new_id_dict[parent_id]
#
#     if first:
#         new_data['cat_id'] = start_index
#         first = False
#     new_cat_id = dao.insert_without_exit("db_category_new", new_data, new_data,'cat_id')
#     old_new_id_dict[old_cat_id] = new_cat_id
#




# part_goods_base_sql = "SELECT * FROM db_monkey_part_goods_base  WHERE part_code = 'C11015100'"
# part_base_array = dao.db.get_data(part_goods_base_sql)
# for part_data in part_base_array:
#     # 复制一份，更改原来的
#     base_id = part_data['id']
#     old_base_uuId = part_data['uuId']
#     update_sql = "UPDATE db_monkey_part_goods_base SET part_name = '后门外拉手饰盖(左)' WHERE id = " + str(base_id)
#     print update_sql
#     dao.db.update_data(update_sql)
#
#     part_base_insert_data = dict(part_data)
#     part_base_insert_data.pop('id')
#     part_base_insert_data.pop('gmt_create')
#     part_base_insert_data.pop('gmt_modified')
#     part_base_insert_data['part_name'] = '后门外拉手饰盖(右)'
#     part_base_uuId = str(uuid.uuid1())
#     part_base_insert_data['uuId'] = part_base_uuId
#
#     dao.insert_temple("db_monkey_part_goods_base", part_base_insert_data)
#     # print part_base_insert_data
#
#     # goods表进行更改
#     part_goods_sql = "SELECT * FROM db_monkey_part_goods WHERE goods_base_id = '" + old_base_uuId + "'"
#     part_goods_array = dao.db.get_data(part_goods_sql)
#     for part_goods_data in part_goods_array:
#         old_part_uuId = part_goods_data['uuId']
#         part_uuId = str(uuid.uuid1())
#         # print part_uuId
#         part_insert_data = part_goods_data
#         part_insert_data.pop('id')
#         part_insert_data.pop('gmt_create')
#         part_insert_data.pop('gmt_modified')
#         part_insert_data['uuId'] = str(part_uuId)
#         part_insert_data['goods_base_id'] = part_base_uuId
#         dao.insert_temple("db_monkey_part_goods", part_insert_data)
#
#         relation_sql = "SELECT * FROM db_monkey_part_liyang_relation WHERE goods_id = '" + str(old_part_uuId) + "'"
#         # print relation_sql
#         relation_array = dao.db.get_data(relation_sql)
#         for relation_data in relation_array:
#             relation_insert_data = relation_data
#
#             relation_insert_data.pop('id')
#             relation_insert_data.pop('gmt_create')
#             relation_insert_data.pop('gmt_modified')
#             relation_insert_data['goods_id'] = part_uuId
#             # print relation_insert_data
#             dao.insert_temple("db_monkey_part_liyang_relation", relation_insert_data)
#



# 写文件
# file_name = r'/Users/zxg/Desktop/temp/cn/test1.xlsx'
# wb = Workbook()
# ew = ExcelWriter(workbook=wb)
# ws = wb.create_sheet(u"基础",0)
# #
# # ws=wb.worksheets[0]#取得wb1的第一个工作表ws1
# # ws.title="socialrange"#指定ws1名字为socialrange
#
# ws.cell(row=1, column=1).value = u'唯一标识'
#
# ew.save(file_name)


# 正则替换
# a = u" 2我3w ew是 rw + 饿w ,.啊@#$%^&*/"
# a1 = re.subn(u'[\u4e00-\u9fa5]+', '', a.strip())
# print a1
# zh_pattern = re.compile(u'[\u4e00-\u9fa5]+')
# match = zh_pattern.search(a)
# if match:
#     print u'oe中有中文：%s' % a
#
# e = "23%$w是我呀"
# b = re.subn(u'[^\w]', '', e)
# print b[0].upper()
#
# oe_string = u"as,v，c"
# oe_array = oe_string.replace("，", ",").split(",")
# print oe_array
'''
# 测试替换乱码
http = HttpUtil.HttpUtil()
result = unicode(http.http_get('http://www.cn357.com/product_list2-1_61-411'), errors='ignore')

need_replace = unicode("�", errors='ignore')
result = result.replace((u'�'), "")
# tree = etree.HTML(result.decode('utf-8'))
# body = tree.xpath("//body")
print type(result)
print result
'''

'''
# 处理center_brand
sql = "select id,name_ch,name_en from center_goods_brand ORDER BY id"
data_array = dao.db.get_data(sql)

for data in data_array:

    id = str(data['id'])
    name_ch = str(data['name_ch'])
    name_en = str(data['name_en'])
    if name_ch == '':
        name = name_en
    elif name_en == '':
        name = name_ch
    else:
        name = name_ch+"/"+name_en

    dao.update_temple("center_goods_brand",{'name':name},{'id':id})
'''



# 处理生成的goods_car关系
'''


# 现有的delete
all_file_object = open('/Users/zxg/Documents/tempFile/tempFile/2015-10-29/15_50_deleteGoodsCar.text')
all_list_of_all_the_lines = list()
try:
    all_list_of_all_the_lines = all_file_object.readlines()
finally:
    all_file_object.close()

for all_line in all_list_of_all_the_lines:
    all_line = all_line.strip('\r\n')
    if all_line == '' or all_line == '\n':
        continue
    if all_line in finally_list:
        continue

    finally_list.append(all_line)


# 错误删除的数据
wrong_delete_list = list()
# 过去历史的delete

one_file_object = open('/Users/zxg/Desktop/test/1.delete_20150918.txt')
one_list_of_all_the_lines = list()
try:
    one_list_of_all_the_lines = one_file_object.readlines()
finally:
    one_file_object.close()
two_file_object = open('/Users/zxg/Desktop/test/2.2015_09_15_17.txt')
two_list_of_all_the_lines = list()
try:
    two_list_of_all_the_lines = two_file_object.readlines()
finally:
    two_file_object.close()

for one_line in one_list_of_all_the_lines:
    print one_line
    one_line = one_line.strip('\n')

    if one_line == '' or one_line == '\n':
        continue
    if one_line in finally_list:
        continue
    finally_list.append(one_line)
    wrong_delete_list.append(one_line)
print 'end one'

for two_line in two_list_of_all_the_lines:
    print two_line
    two_line = two_line.strip('\n')

    if two_line == '' or two_line == '\n':
        continue
    if two_line in finally_list:
        continue
    finally_list.append(two_line)
    wrong_delete_list.append(two_line)

print 'end two'

print len(one_list_of_all_the_lines)
print len(two_list_of_all_the_lines)
print len(finally_list)




# 写入
file_object = open('/Users/zxg/Desktop/test/delete.txt', 'w')
try:
    file_object.writelines("\n".join(finally_list))
finally:
    file_object.close()

# 写入
wrong_file_object = open('/Users/zxg/Desktop/test/wrongdelete.txt', 'w')
try:
    wrong_file_object.writelines("\n".join(wrong_delete_list))
finally:
    wrong_file_object.clos
()

'''
