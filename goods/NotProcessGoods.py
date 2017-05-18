# encoding=utf-8
# 没有处理的商品
import os

__author__ = 'zxg'


from openpyxl import Workbook
from openpyxl.writer.excel import ExcelWriter

from util import CrawlDao

athena_dao = CrawlDao.CrawlDao("dataserver","online_cong")


goods_car_list = list()
goods_car_array = athena_dao.db.get_data("select goods_id from db_goods_car where status = 1 group by goods_id")
for goods_car_data in goods_car_array:
    goods_car_list.append(str(goods_car_data['goods_id']))



stall_dao = CrawlDao.CrawlDao("modeldatas","test")
goods_array = stall_dao.db.get_data("select dg.goods_id as goods_id ,dg.goods_name as goods_name,db.brand_name as brand_name,dg.goods_format  as goods_format  "
                                    "from db_goods dg,db_brand db "
                                    "where dg.brand_id = db.brand_id "
                                    "and dg.data_flag = 0 "
                                    "and dg.is_delete = 0  "
                                    "and dg.seller_id = 1")



wb = Workbook()
ew = ExcelWriter(workbook=wb)

have_sheet = wb.create_sheet(u"有车型", 0)
not_have_sheet = wb.create_sheet(u"无车型", 0)

have_sheet.cell(row=1, column=1).value = 'goods_id'
have_sheet.cell(row=1, column=2).value = '名称'
have_sheet.cell(row=1, column=3).value = '品牌'
have_sheet.cell(row=1, column=4).value = '规格'
not_have_sheet.cell(row=1, column=1).value = 'goods_id'
not_have_sheet.cell(row=1, column=2).value = '名称'
not_have_sheet.cell(row=1, column=3).value = '品牌'
not_have_sheet.cell(row=1, column=4).value = '规格'

have_row_number = 2
not_row_number = 2

for goods_data in goods_array:
    goods_id = str(goods_data['goods_id'])
    goods_name = str(goods_data['goods_name'])
    brand_name = str(goods_data['brand_name'])
    goods_format = str(goods_data['goods_format'])

    if goods_id in goods_car_list:
        have_sheet.cell(row=have_row_number, column=1).value = goods_id
        have_sheet.cell(row=have_row_number, column=2).value = goods_name
        have_sheet.cell(row=have_row_number, column=3).value = brand_name
        have_sheet.cell(row=have_row_number, column=4).value = goods_format
        have_row_number += 1
    else:
        not_have_sheet.cell(row=not_row_number, column=1).value = goods_id
        not_have_sheet.cell(row=not_row_number, column=2).value = goods_name
        not_have_sheet.cell(row=not_row_number, column=3).value = brand_name
        not_have_sheet.cell(row=not_row_number, column=4).value = goods_format
        not_row_number += 1

ew.save(filename=os.getcwd() + '/待整理的商品.xlsx')




