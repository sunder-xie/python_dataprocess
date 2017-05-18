# encoding=utf-8
# just test
# 导出当前品牌的数据，分brand的sheet导出excel
import os

__author__ = 'zxg'

import sys

from openpyxl import Workbook
from openpyxl.writer.excel import ExcelWriter

reload(sys)
sys.setdefaultencoding("utf-8")
# sys.path.append(os.getcwd() + "../util/")

from util import CrawlDao


class ExportExcelGoods:
    def __init__(self):
        self.dao = CrawlDao.CrawlDao("modeldatas", "online_cong")
        # self.dao = CrawlDao.CrawlDao("modeldatas", "local")

    def create_sheet(self, the_sheet):
        the_sheet.cell(row=1, column=1).value = u'商品名称'
        the_sheet.cell(row=1, column=2).value = u'OE码'
        the_sheet.cell(row=1, column=3).value = u'规格型号'
        the_sheet.cell(row=1, column=4).value = u'是否为通用件'
        the_sheet.cell(row=1, column=5).value = u'车型别称'
        the_sheet.cell(row=1, column=6).value = u'商品属性'
        the_sheet.cell(row=1, column=7).value = u'使用车类型'
        the_sheet.cell(row=1, column=8).value = u'档口价'
        the_sheet.cell(row=1, column=9).value = u'保险价'
        the_sheet.cell(row=1, column=10).value = u'城市'
        the_sheet.cell(row=1, column=11).value = u'库存值'
        the_sheet.cell(row=1, column=12).value = u'单位描述'
        the_sheet.cell(row=1, column=13).value = u'是否需要保质期管理'
        the_sheet.cell(row=1, column=14).value = u'保质期时间'
        the_sheet.cell(row=1, column=15).value = u'预警时限'

    def main(self, brand_name, save_file):
        wb = Workbook()
        ew = ExcelWriter(workbook=wb)

        relation_db_name = str(self.dao.db.get_data("select liyang_table from db_monkey_part_liyang_table_relation where car_brand_name = '"+brand_name+"'" )[0]['liyang_table'])

        model_array = self.dao.db.get_data("select id,liyang_factory,liyang_model,liyang_series "
                                           "from db_monkey_part_liyang_base "
                                           "where liyang_brand = '" + brand_name + "' and is_deleted = 'N'")
        sheet_index = 0
        for model_data in model_array:
            id = str(model_data['id'])
            # create sheet
            sheet_name = str(model_data['liyang_factory']) + "-" + str(model_data['liyang_series']) + "-" + str(
                model_data['liyang_model'])
            the_sheet = wb.create_sheet(sheet_name.decode("utf-8"), sheet_index)
            sheet_index += 1

            self.create_sheet(the_sheet)

            row_num = 1
            goods_sql = "select dg.part_name as part_name,dg.oe_number as oe from " \
                        "(select goods_id " \
                        "from "+relation_db_name+" " \
                        "where part_liyang_id = " + id + \
                        " group by goods_id " \
                        ") dmplr,db_monkey_part_goods_base dg  " \
                        "where dg.uuid = dmplr.goods_id "
            goods_array = self.dao.db.get_data(goods_sql)
            for goods_data in goods_array:
                row_num += 1
                the_sheet.cell(row=row_num, column=1).value = str(goods_data['part_name'])
                the_sheet.cell(row=row_num, column=2).value = str(goods_data['oe'])

        ew.save(filename=save_file)


brand_name = '现代'
save_file = os.getcwd() + "/" + brand_name + ".xlsx"

exportExcelGoods = ExportExcelGoods()
exportExcelGoods.main(brand_name, save_file)
