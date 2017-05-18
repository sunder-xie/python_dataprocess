# encoding=utf-8
# 数据平台新增的分类，生成 至电商库、中心商品库的分类数据
# 04.05 更新--待测试
import datetime
import os

__author__ = 'zxg'

from util import CrawlDao

from util.pinying.pinyin import PinYin
from openpyxl import Workbook
from openpyxl.writer.excel import ExcelWriter


class MonkeyToOther:
    def __init__(self):
        self.monkey_dao = CrawlDao.CrawlDao("modeldatas", "local")
        self.stall_dao = CrawlDao.CrawlDao("autoparts", "local")

        self.pinyinDO = PinYin()
        self.pinyinDO.load_word()

        self.stall_insert_data_list = list()

        # self.the_time = '2016-08'
        self.the_time = datetime.datetime.now().strftime("%Y-%m")

        # 不同的cat_code长度确定层级
        self.cat_level_by_code = {2: 1, 5: 2, 9: 3}
        # 新增的分类
        monkey_cat_dict = dict()
        monkey_cat_array = self.monkey_dao.db.get_data(
            "select cat_id,parent_id,cat_code from db_category where is_deleted = 'N'")
        for monkey_cat_data in monkey_cat_array:
            monkey_cat_dict[str(monkey_cat_data['cat_id'])] = monkey_cat_data

        self.new_cat_list = list()
        new_cat_array = self.monkey_dao.db.get_data(
            "select cat_name,parent_id,cat_level,cat_code,vehicle_code from db_category where is_deleted = 'N' and gmt_create like '" + self.the_time + "%' ")
        for new_cat_data in new_cat_array:
            cat_code = str(new_cat_data['cat_code'])
            parent_id = str(new_cat_data['parent_id'])
            true_cat_code = self.get_all_code(cat_code, parent_id, monkey_cat_dict)

            new_data = {
                'cat_name': str(new_cat_data['cat_name']),
                'vehicle_code': str(new_cat_data['vehicle_code']),
                'cat_code': true_cat_code
            }
            self.new_cat_list.append(new_data)

        # stall
        self.stall_cat_id = 0
        self.stall_cat_dict = dict()
        stall_cat_array = self.stall_dao.db.get_data(
            "select cat_id,cat_code,vehicle_code from db_category where is_deleted = 'N' ")
        for stall_data in stall_cat_array:
            self.stall_cat_dict[str(stall_data['cat_code'])] = stall_data
            cat_id = int(stall_data['cat_id'])
            if cat_id > self.stall_cat_id:
                self.stall_cat_id = cat_id

    def to_stall(self):
        for new_cat_data in self.new_cat_list:
            cat_name = new_cat_data['cat_name']
            vehicle_code = new_cat_data['vehicle_code']
            if cat_name == '怠速电机':
                continue
            cat_code = str(new_cat_data['cat_code'])
            stall_parent_id = self.get_parent_id(cat_code)

            cat_code_array = cat_code.split(".")
            sort_order = cat_code_array[len(cat_code_array) - 1]
            if len(cat_code) != 9 or vehicle_code != "CH":
                if len(cat_code) != 9:
                    print '非三级catcode：%s ' % cat_code
                else:
                    print '三级 为单一属性:%s' % cat_code
                self.stall_cat_id += 1
                save_data = {
                    'cat_id': self.stall_cat_id,
                    'cat_name': cat_name,
                    'parent_id': stall_parent_id,
                    'sort_order': sort_order,
                    'cat_level': self.cat_level_by_code[len(cat_code)],
                    'cat_code': cat_code,
                    'vehicle_code': vehicle_code
                }
                self.stall_cat_dict[cat_code] = save_data
                self.stall_insert_data_list.append(save_data)
            else:
                for ve in ['C', 'H']:
                    self.stall_cat_id += 1
                    save_data = {
                        'cat_name': cat_name,
                        'parent_id': stall_parent_id,
                        'sort_order': sort_order,
                        'cat_level': '3',
                        'cat_code': cat_code,
                        'vehicle_code': ve
                    }
                    self.stall_cat_dict[cat_code] = save_data
                    self.stall_insert_data_list.append(save_data)

    def write_cat_sql(self, stall_address, center_address):
        stall_sql_list = list()
        center_sql_list = list()
        for save_data in self.stall_insert_data_list:
            stall_sql = self.get_insert_sql("db_category", save_data)
            stall_sql_list.append(stall_sql)

            cat_name = save_data['cat_name']
            first_letter = self.pinyinDO.firstLetter(cat_name)
            save_data['first_letter'] = first_letter
            save_data.pop("cat_id")
            center_sql = self.get_insert_sql("center_category", save_data)
            center_sql_list.append(center_sql)

        stall_file_object = open(stall_address, 'w')
        try:
            stall_file_object.writelines(";\n".join(stall_sql_list))
        finally:
            stall_file_object.close()

        center_file_object = open(center_address, 'w')
        try:
            center_file_object.writelines(";\n".join(center_sql_list))
        finally:
            center_file_object.close()

    def write_part_center_sql(self, center_part_address):
        part_sql_list = list()
        # 新增的标准零件名称
        new_part_array = self.monkey_dao.db.get_data(
            "select part_name,part_code,sum_code,aliss_name_text,label_name_text,cat_kind,vehicle_code from db_category_part where is_deleted = 'N' and gmt_create like '" + self.the_time + "%' ")
        for part_data in new_part_array:
            part_sql = self.get_insert_sql("center_part", part_data)
            part_sql_list.append(part_sql)

        center_file_object = open(center_part_address, 'w')
        try:
            center_file_object.writelines(";\n".join(part_sql_list))
        finally:
            center_file_object.close()

    def write_jindie_excel(self, jin_die_address):
        wb = Workbook()
        ew = ExcelWriter(workbook=wb)

        cat_sheet = wb.create_sheet(u"1", 0)

        exist_list = list()
        row_num = 0
        for save_data in self.stall_insert_data_list:
            cat_code = str(save_data['cat_code'])
            cat_name = str(save_data['cat_name'])

            key = cat_code + "-" + cat_name
            if key not in exist_list:
                row_num += 1
                exist_list.append(key)
                cat_sheet.cell(row=row_num, column=1).value = cat_code
                cat_sheet.cell(row=row_num, column=2).value = cat_name
                cat_sheet.cell(row=row_num, column=3).value = u'FALSE'

        ew.save(filename=jin_die_address)

    def main_def(self, stall_cat_address, center_cat_address, center_part_address, jin_die_address):
        self.to_stall()
        self.write_cat_sql(stall_cat_address, center_cat_address)
        self.write_part_center_sql(center_part_address)
        self.write_jindie_excel(jin_die_address)

    # ============= 辅助执行类======================================

    # 根据cat_code获得其父id
    def get_parent_id(self, cat_code):
        code_length = len(cat_code)

        if code_length == 9:
            parent_code = cat_code[:5]
        elif code_length == 5:
            parent_code = cat_code[:2]
        else:
            # 一级分类
            return '0'

        parent_data = self.stall_cat_dict[parent_code]

        return str(parent_data['cat_id'])

    # monkey 获得其全的code
    def get_all_code(self, cat_code, parent_id, monkey_cat_dict=dict()):
        if parent_id == "0":
            return cat_code
        else:
            parent_data = monkey_cat_dict[parent_id]
            parent_code = str(parent_data['cat_code'])

            cat_code = parent_code + "." + cat_code

            return self.get_all_code(cat_code, str(parent_data['parent_id']), monkey_cat_dict)

    # 拼写成sql
    def get_insert_sql(self, table, dic):
        gmt = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        dic['gmt_create'] = gmt
        dic['gmt_modified'] = gmt
        sql = 'insert ignore into ' + table + '(' + ','.join(dic.keys()) + ') values'
        value_list = list()
        for key, value in dic.items():
            value = str(value).replace('"', '')
            value_list.append('"' + self.monkey_dao.html_tag.sub('', str(value)) + '"')
        sql += '(' + ','.join(value_list) + ')'
        return sql


the_time = datetime.datetime.now().strftime("%Y-%m-%d-%H")
stall_address = os.getcwd() + "/sql/"+the_time+"stall_cat_insert.sql"
center_address = os.getcwd() + "/sql/"+the_time+"center_cat_insert.sql"
center_part_address = os.getcwd() + "/sql/"+the_time+"center_part_insert.sql"
jin_die_address = os.getcwd() + "/sql/"+the_time+"new_jindie.xlsx"

monkeyToCenter = MonkeyToOther()
monkeyToCenter.main_def(stall_address, center_address, center_part_address, jin_die_address)
