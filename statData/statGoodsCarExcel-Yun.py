# encoding=utf-8
# 统计云修excel商品占比 2016.09.12
import copy
import json

__author__ = 'zxg'


# from __future__ import division  #除法返回真实的商

from openpyxl import Workbook
from openpyxl.writer.excel import ExcelWriter
from util import CrawlDao, FileUtil


class GoodsCarExcel:
    def __init__(self):
        self.dao = CrawlDao.CrawlDao('test', "local")
        self.fileDao = FileUtil.FileDao()

        self.file_parent = r'/Users/zxg/Desktop/yunxiulvqing'
        # ===初始化的变量====
        # brand_factory:list<map>{leyel_id,series,vehicle_type,model_year,displacement ,intake_style ,	max_power ,	fuel_type}
        self.liyang_dict_key = "{}_{}_{}"
        self.liyang_dict = dict()
        # liyang_id:online_car_id
        self.liyang_tqmall_dict = dict()

        # car_category
        # online_car_id_list
        self.car_category_list = list()
        # id:{brand,series,model,power,year,name}
        self.car_category_dict = dict()

        self.init_liyang()
        self.init_liyang_tqmall()
        self.init_car()

        # ====处理excel的变量===
        # 云修id:set(online_car_id)
        self.goods_car_dict = dict()

        # online_car_id:set(云修id)
        self.car_goods_dict = dict()
        # liyang_car_id:set(云修id)
        self.liyang_car_goods_dict = dict()
        # liyang_id:{liyang_data}
        self.liyang_id_data_dict = dict()

        # 不匹配的车型列表
        self.not_car_list = copy.deepcopy(self.car_category_list)

        # excle中没有匹配上的数据
        self.wrong_data_list = list()

    # ========start 初始化===============
    # 初始化db_car_info_all 数据
    def init_liyang(self):
        liyang_sql = "select leyel_id,factory_name ,car_brand ,car_series ,	vehicle_type ,	model_year ," \
                     "displacement ,intake_style ,	max_power ,	fuel_type,create_year  from db_car_info_all"
        liyang_array = self.dao.db.get_data(liyang_sql)
        for liyang_data in liyang_array:
            factory_name = str(liyang_data['factory_name'])
            car_brand = str(liyang_data['car_brand'])
            car_series = str(liyang_data['car_series'])

            liyang_car_data = {
                "leyel_id": str(liyang_data['leyel_id']),
                "vehicle_type": str(liyang_data['vehicle_type']),
                "model_year": str(liyang_data['model_year']),
                "displacement": str(liyang_data['displacement']),
                "intake_style": str(liyang_data['intake_style']),
                "max_power": str(liyang_data['max_power']),
                "fuel_type": str(liyang_data['fuel_type']),
                "create_year": str(liyang_data['create_year'])
            }

            key = self.liyang_dict_key.format(car_brand, factory_name, car_series)
            if key in self.liyang_dict.keys():
                liyang_car_list = list(self.liyang_dict[key])
            else:
                liyang_car_list = list()
            liyang_car_list.append(liyang_car_data)
            self.liyang_dict[key] = liyang_car_list

    # 初始化db_car_all 数据
    def init_liyang_tqmall(self):
        car_all_sql = "select new_l_id,car_models_id from db_car_all"
        car_all_array = self.dao.db.get_data(car_all_sql)
        for car_all_data in car_all_array:
            l_id = str(car_all_data['new_l_id'])
            self.liyang_tqmall_dict[l_id] = str(car_all_data['car_models_id'])

    # db_car_category 数据
    def init_car(self):
        car_categort = "select id,brand,series,model,power,year,name from db_car_category where level = 6 and is_del = 0"
        car_category_array = self.dao.db.get_data(car_categort)
        for car_category_data in car_category_array:
            id = str(car_category_data["id"])
            self.car_category_list.append(id)
            self.car_category_dict[id] = car_category_data

    # ========end 初始化===============

    # ========start 数据处理===============
    # 判断这个数据是否符合liyang原始数据的标准
    # leyel_id,vehicle_type,model_year,displacement ,intake_style ,	max_power ,	fuel_type
    def find_true_liyang(self, liyang_data, car_model, car_year, car_display, car_intake, car_power, car_flue):
        # print 'liyang_data:%s.' % json.dumps(liyang_data)
        #
        if car_model != "":
            car_model_array = car_model.split("/")
            is_not_exist = True
            for car_model_name in car_model_array:
                if car_model_name == liyang_data['vehicle_type']:
                    is_not_exist = False
            if is_not_exist:
                # print 'wrong car_model:%s.' % car_model_name
                return False
        if car_year != "":
            if "-" in car_year:
                car_year_array = car_year.split("-")
                start_year = car_year_array[0]
                end_year = car_year_array[1]
            else:
                start_year = car_year
                end_year = car_year
            liyang_year = int(liyang_data['model_year'])
            if (start_year != "" and int(start_year) > liyang_year) or (end_year != "" and int(end_year) < liyang_year):
                # print 'wrong liyang_year:%s.' % liyang_year
                return False

        if car_display != "":
            car_display_array = car_display.split("/")
            is_not_exist = True
            for car_display_name in car_display_array:
                if car_display_name == liyang_data['displacement']:
                    is_not_exist = False
            if is_not_exist:
                # print 'wrong car_display_name:%s.' % car_display_name

                return False
        if car_intake != "":
            car_intake_array = car_intake.split("/")
            is_not_exist = True
            for car_intake_name in car_intake_array:
                if car_intake_name == liyang_data['intake_style']:
                    is_not_exist = False
            if is_not_exist:
                # print 'wrong car_intake_name:%s.' % car_intake_name

                return False
        if car_power != "":
            # car_power_int = float(car_power)
            # max_power = liyang_data['max_power']
            if car_power not in liyang_data['max_power']:
                is_not_exist = True
                # 小数 97.8 ~~ 98
                # if "/" not in max_power:
                #     max_power_array = max_power.split(";")
                #     for max_power_va in max_power_array:
                #         if max_power_va == "":
                #             continue
                #         max_power_value = float(max_power_va)
                #         if car_power_int > max_power_value > (max_power_value - 1.0):
                #             is_not_exist = False

                # print 'wrong car_power:%s.' % car_power
                if is_not_exist:
                    return False
        if car_flue != "":
            if car_flue != liyang_data['fuel_type']:
                # print 'wrong car_flue:%s.' % car_flue

                return False

        return True

    def read_excel(self, file_address):
        # 单个excle处理
        data = self.fileDao.open_excel(file_address)
        #
        table = data.sheets()[0]

        n_rows = table.nrows  # 行数
        n_cols = table.ncols  # 列数
        print ('行数：%s ,列数：%s' % (n_rows, n_cols))

        liyang_tqmall_dict_keys = self.liyang_tqmall_dict.keys()
        liyang_dict_keys = self.liyang_dict.keys()
        for rownum in range(1, n_rows):
            row = table.row_values(rownum)
            goods_format = str(row[5]).strip()
            goods_size = str(row[6]).strip()
            car_factory = str(row[7]).strip().replace(".0", "")
            car_brand = str(row[8]).strip().replace(".0", "")
            car_series = str(row[9]).strip().replace(".0", "")
            car_model = str(row[10]).strip().replace(".0", "")
            car_year = str(row[11]).strip().replace(".0", "")
            car_display = str(row[12]).strip()
            car_intake = str(row[13]).strip()
            car_power = str(row[14]).strip().replace(".0", "")
            car_flue = str(row[15]).strip()
            yun_id = str(row[16]).strip()


            if yun_id in self.goods_car_dict.keys():
                car_id_set = set(self.goods_car_dict[yun_id])
            else:
                car_id_set = set()

            key = self.liyang_dict_key.format(car_brand, car_factory, car_series)
            if key not in liyang_dict_keys:
                wrong_str = "{} not in liyang_dict".format(key)
                self.wrong_data_list.append(wrong_str)
                print wrong_str
                continue

            try:
                liyang_data_list = self.liyang_dict.get(key)
            except Exception, e:
                print 'self.liyang_dict.get(key) error:' + key
                print e
                continue

            # 是否对应到具体的数据
            is_find = False
            for liyang_data in liyang_data_list:
                is_ok_liyang = self.find_true_liyang(liyang_data, car_model, car_year, car_display, car_intake,
                                                     car_power, car_flue)
                if is_ok_liyang:
                    is_find = True
                    liyang_id = liyang_data['leyel_id']

                    if liyang_id not in liyang_tqmall_dict_keys:
                        wrong_str = "liyang_id not in db_car_all,not find online_car_id,liyang_id:" + liyang_id
                        self.wrong_data_list.append(wrong_str)
                        print wrong_str
                        continue

                    online_car_id = self.liyang_tqmall_dict[liyang_id]
                    # 添加car对应的goods列表
                    if online_car_id in self.car_goods_dict.keys():
                        goods_set = set(self.car_goods_dict[online_car_id])
                    else:
                        goods_set = set()
                    goods_set.add(yun_id)
                    self.car_goods_dict[online_car_id] = goods_set

                    if liyang_id in self.liyang_car_goods_dict.keys():
                        liyang_goods_set = set(self.liyang_car_goods_dict[liyang_id])
                    else:
                        liyang_goods_set = set()
                    liyang_goods_set.add(yun_id)
                    self.liyang_car_goods_dict[liyang_id] = liyang_goods_set

                    liyang_id_data = liyang_data
                    liyang_id_data['factory_name'] = car_factory
                    liyang_id_data['car_brand'] = car_brand
                    liyang_id_data['car_series'] = car_series
                    self.liyang_id_data_dict[liyang_id] = liyang_id_data

                    # 添加goods 对应的car 列表
                    car_id_set.add(online_car_id)
                self.goods_car_dict[yun_id] = car_id_set

            # 没有对应到
            if not is_find:
                wrong_str = 'not liyang car.yun_id:{},brand:{}, factory:{}, series:{},model:{},year:{},display:{},intake:{},power:{},燃料类型:{}'.format(
                    yun_id, car_brand, car_factory, car_series, car_model, car_year, car_display, car_intake,
                    car_power, car_flue)
                print wrong_str
                self.wrong_data_list.append(wrong_str)

    def read_special_excel(self, file_address):
        print '======== 特殊的excel 处理 ========'
        # 单个excle处理
        data = self.fileDao.open_excel(file_address)

        #
        table = data.sheets()[0]

        n_rows = table.nrows  # 行数
        n_cols = table.ncols  # 列数
        print ('行数：%s ,列数：%s' % (n_rows, n_cols))

        liyang_tqmall_dict_keys = self.liyang_tqmall_dict.keys()
        liyang_dict_keys = self.liyang_dict.keys()
        for rownum in range(1, n_rows):
            row = table.row_values(rownum)
            car_factory = str(row[6]).strip().replace(".0", "")
            car_brand = str(row[7]).strip().replace(".0", "")
            car_series = str(row[8]).strip().replace(".0", "")
            car_model = str(row[9]).strip().replace(".0", "")

            car_display = str(row[10]).strip()
            car_intake = str(row[11]).strip()
            car_create_year = str(row[12]).strip()
            car_power = str(row[13]).strip().replace(".0", "")

            yun_id = str(row[14]).strip()


            if yun_id in self.goods_car_dict.keys():
                car_id_set = set(self.goods_car_dict[yun_id])
            else:
                car_id_set = set()

            key = self.liyang_dict_key.format(car_brand, car_factory, car_series)
            if key not in liyang_dict_keys:
                wrong_str = "{} not in liyang_dict".format(key)
                self.wrong_data_list.append(wrong_str)
                print wrong_str
                continue

            try:
                liyang_data_list = self.liyang_dict.get(key)
            except Exception, e:
                print 'self.liyang_dict.get(key) error:' + key
                print e
                continue

            # 是否对应到具体的数据
            is_find = False
            for liyang_data in liyang_data_list:
                is_ok_liyang = self.find_true_liyang(liyang_data, car_model, car_year, car_display, car_intake,
                                                     car_power, car_flue)
                if is_ok_liyang:
                    is_find = True
                    liyang_id = liyang_data['leyel_id']

                    if liyang_id not in liyang_tqmall_dict_keys:
                        wrong_str = "liyang_id not in db_car_all,not find online_car_id,liyang_id:" + liyang_id
                        self.wrong_data_list.append(wrong_str)
                        print wrong_str
                        continue

                    online_car_id = self.liyang_tqmall_dict[liyang_id]
                    # 添加car对应的goods列表
                    if online_car_id in self.car_goods_dict.keys():
                        goods_set = set(self.car_goods_dict[online_car_id])
                    else:
                        goods_set = set()
                    goods_set.add(yun_id)
                    self.car_goods_dict[online_car_id] = goods_set

                    if liyang_id in self.liyang_car_goods_dict.keys():
                        liyang_goods_set = set(self.liyang_car_goods_dict[liyang_id])
                    else:
                        liyang_goods_set = set()
                    liyang_goods_set.add(yun_id)
                    self.liyang_car_goods_dict[liyang_id] = liyang_goods_set

                    liyang_id_data = liyang_data
                    liyang_id_data['factory_name'] = car_factory
                    liyang_id_data['car_brand'] = car_brand
                    liyang_id_data['car_series'] = car_series
                    self.liyang_id_data_dict[liyang_id] = liyang_id_data

                    # 添加goods 对应的car 列表
                    car_id_set.add(online_car_id)
                self.goods_car_dict[yun_id] = car_id_set

            # 没有对应到
            if not is_find:
                wrong_str = 'not liyang car.yun_id:{},brand:{}, factory:{}, series:{},model:{},year:{},display:{},intake:{},power:{},燃料类型:{}'.format(
                    yun_id, car_brand, car_factory, car_series, car_model, car_year, car_display, car_intake,
                    car_power, car_flue)
                print wrong_str
                self.wrong_data_list.append(wrong_str)

    def write_sku_excel(self):
        wb = Workbook()
        ew = ExcelWriter(workbook=wb)
        the_sheet = wb.create_sheet(u"sku适配车型", 0)
        insert_list = (u'云修号', u'car_id', u'品牌', u'车系', u'车型', u'排量', u'年款', u'名称')
        self.write_excel_cell(the_sheet, 1, insert_list)

        is_sheet_num = 2
        for yun_id, car_id_set in self.goods_car_dict.iteritems():
            if len(car_id_set) == 0:
                self.write_excel_cell(the_sheet, is_sheet_num, (
                    yun_id, u'None', u'None', u'None', u'None', u'None', u'None', u'None'))
                is_sheet_num += 1
            else:
                for car_id in car_id_set:
                    if car_id in self.not_car_list:
                        self.not_car_list.remove(car_id)

                    # id,brand,series,model,power,year,name
                    goods_car_data = self.car_category_dict[car_id]
                    self.write_excel_cell(the_sheet, is_sheet_num, (
                        yun_id, str(car_id), str(goods_car_data['brand']),
                        str(goods_car_data['series']), str(goods_car_data['model']),
                        str(goods_car_data['power']), str(goods_car_data['year']), str(goods_car_data['name'])))
                    is_sheet_num += 1

        ew.save(filename=self.file_parent + '/1.滤清器适配车型.xlsx')

    def write_not_car_excel(self):
        wb = Workbook()
        ew = ExcelWriter(workbook=wb)
        the_sheet = wb.create_sheet(u"sku不适配车型", 0)
        insert_list = (u'car_id', u'品牌', u'车系', u'车型', u'排量', u'年款', u'名称')
        self.write_excel_cell(the_sheet, 1, insert_list)
        is_sheet_num = 2
        for car_id in self.not_car_list:
            goods_car_data = self.car_category_dict[car_id]
            self.write_excel_cell(the_sheet, is_sheet_num, (
                str(car_id), str(goods_car_data['brand']),
                str(goods_car_data['series']), str(goods_car_data['model']),
                str(goods_car_data['power']), str(goods_car_data['year']), str(goods_car_data['name'])))
            is_sheet_num += 1
        ew.save(filename=self.file_parent + '/2.滤清器不适配车型.xlsx')

    def write_1_n_excel(self):
        wb = Workbook()
        ew = ExcelWriter(workbook=wb)
        the_sheet = wb.create_sheet(u"sku不适配车型", 0)
        insert_list = (u'car_id', u'品牌', u'车系', u'车型', u'排量', u'年款', u'名称', u'云修号')
        self.write_excel_cell(the_sheet, 1, insert_list)
        is_sheet_num = 2
        for car_id, goods_set in self.car_goods_dict.iteritems():
            if len(goods_set) < 2:
                continue
            goods_car_data = self.car_category_dict[car_id]
            for yun_id in goods_set:
                self.write_excel_cell(the_sheet, is_sheet_num, (
                    str(car_id), str(goods_car_data['brand']),
                    str(goods_car_data['series']), str(goods_car_data['model']),
                    str(goods_car_data['power']), str(goods_car_data['year']), str(goods_car_data['name']),
                    yun_id))
                is_sheet_num += 1

        ew.save(filename=self.file_parent + '/3.淘汽车型对应多个机滤.xlsx')

    def write_1_n_liyang_excel(self):
        wb = Workbook()
        ew = ExcelWriter(workbook=wb)
        the_sheet = wb.create_sheet(u"sku不适配车型", 0)
        insert_list = (u'liyang_id', u'品牌', u'厂家', u'车系', u'车型', u'年款',u'生产年份', u'排量', u'进气形式', u'最大功率', u'燃料类型', u'云修号')
        self.write_excel_cell(the_sheet, 1, insert_list)
        is_sheet_num = 2
        for liyang_id, goods_set in self.liyang_car_goods_dict.iteritems():
            if len(goods_set) < 2:
                continue
            liyang_data = self.liyang_id_data_dict[liyang_id]
            for yun_id in goods_set:
                self.write_excel_cell(the_sheet, is_sheet_num, (
                    liyang_id, liyang_data['car_brand'], liyang_data['factory_name'], liyang_data['car_series'],
                    liyang_data['vehicle_type'], liyang_data['model_year'], liyang_data['create_year'], liyang_data['displacement'],
                    liyang_data['intake_style'],
                    liyang_data['max_power'], liyang_data['fuel_type'], yun_id))
                is_sheet_num += 1

        ew.save(filename=self.file_parent + '/4.liyang车型对应多个机滤.xlsx')

    def save_wrong_data(self):
        file_object = open(self.file_parent + "/5.wrong.txt", 'w')

        try:
            file_object.writelines("\n".join(self.wrong_data_list))
        finally:
            file_object.close()

    def main_do(self, console_excel):
        # 正常的excel 处理
        self.read_excel(console_excel)
        # 特殊的云修号的 处理
        # self.read_special_excel(console__special_excel)


        self.write_sku_excel()
        self.write_not_car_excel()
        self.write_1_n_excel()
        self.write_1_n_liyang_excel()
        self.save_wrong_data()

    # ===help

    def write_excel_cell(self, sheet, row_num, value_list):
        start_colum = 1
        for value in value_list:

            try:
                sheet.cell(row=row_num, column=start_colum).value = str(value.decode('UTF-8'))
            except:
                print "error:==== " + str(row_num) + ":" + str(start_colum) + ":=== " + ",".join(value_list)
            start_colum += 1


console_excel = r'/Users/zxg/Documents/work/淘气档口/work/0912机滤匹配/机滤正确数据总表20160913(4).xlsx'
# console__special_excel = r'/Users/zxg/Documents/work/淘气档口/work/0912机滤匹配/云修机滤型号车型匹配-pxf(1)修改版(1)(1).xlsx'
goodsCarExcel = GoodsCarExcel()
goodsCarExcel.main_do(console_excel)
