# encoding=utf-8
# 统计云修牌商品占比 2016.08.23
import copy

__author__ = 'zxg'


# from __future__ import division  #除法返回真实的商

from openpyxl import Workbook
from openpyxl.writer.excel import kj
from util import CrawlDao

class GoodsCarPercentage:
    def __init__(self):
        # self.dao = CrawlDao.CrawlDao('test', "local")
        self.dao = CrawlDao.CrawlDao('dataserver', "online_cong")
        self.shop_dao = CrawlDao.CrawlDao('ol_autoparts', "stall")
        # car_category
        self.car_category_list = list()
        self.car_category_dict = dict()

        # cat:(car_id....)
        self.cat_car_dict = dict()

        # cat_id:cat_name
        self.cat_dict =dict()
        # goods brand
        self.brand_id = 849
        self.brand_sum_num = 0
        self.car_sum_num = 0
        # save filepath's parent
        # self.file_parent = r'/Users/zxg/Desktop/statyunxiu'
        self.file_parent = r'/home/tqmall.pr/statyunxiu'
        self.init_data()

    def init_data(self):
        self.brand_sum_num = int((self.dao.db.get_data("select count(1) as sum from db_car_category where level = 1 and is_del = 0"))[0]['sum'])

        car_categort ="select id,brand,series,model,power,year,name from db_car_category where level = 6 and is_del = 0"
        car_category_array = self.dao.db.get_data(car_categort)
        self.car_sum_num = len(car_category_array)
        for car_category_data in car_category_array:
            id = str(car_category_data["id"])
            self.car_category_list.append(id)
            self.car_category_dict[id] = car_category_data



        cat_sql = "select cat_id,cat_name from db_category where cat_level = 3 and is_deleted = 'N'"
        cat_array = self.shop_dao.db.get_data(cat_sql)
        for cat_data in cat_array:
            self.cat_dict[str(cat_data['cat_id'])] = str(cat_data['cat_name'])


    def write_excel_cell(self,sheet,row_num,value_list):
        start_colum = 1
        for value in value_list:

            try:
                sheet.cell(row=row_num, column=start_colum).value = str(value.decode('UTF-8'))
            except:
                print "error:==== "+str(row_num)+":"+str(start_colum)+":=== "+",".join(value_list)
            start_colum += 1


    # 统计单个 sku
    def stat_sku(self,goods_id,goods_name,cat_id):
        # start get 数据
        goods_car_sql = "select car_id,car_brand,car_series,car_model,car_power,car_year,car_name from db_goods_car  where goods_id ="+goods_id
        goods_car_array = self.dao.db.get_data(goods_car_sql)
        if len(goods_car_array) == 0:
            print "======%s goods car is empty=====" % goods_id
            return False

        if cat_id in self.cat_car_dict.keys():
            cat_car_list = list(self.cat_car_dict[cat_id])
        else:
            cat_car_list = list()

        wb = Workbook()
        ew = ExcelWriter(workbook=wb)
        name_sheet = wb.create_sheet(u"name", 0)
        is_sheet = wb.create_sheet(u"适配车型", 1)
        not_sheet = wb.create_sheet(u"不适配车型", 2)
        is_sheet_num = 2
        not_sheet_num = 2

        name_list = list()
        name_list.append(goods_name)
        self.write_excel_cell(name_sheet,1,name_list)
        # 初始化 sheet
        for the_sheet in (is_sheet,not_sheet):
            insert_list = (u'car_id',u'品牌',u'车系',u'车型',u'排量',u'年款',u'名称')
            self.write_excel_cell(the_sheet,1,insert_list)

        now_all_car_list = copy.deepcopy(self.car_category_list)
        for goods_car_data in goods_car_array:
            car_id = str(goods_car_data['car_id'])
            try:
                now_all_car_list.remove(car_id)
            except:
                print 'error:==== goods_id:%s,car_id:%s' % (goods_id,car_id)
            if car_id not in cat_car_list:
                cat_car_list.append(car_id)

            self.write_excel_cell(is_sheet,is_sheet_num,(str(car_id),str(goods_car_data['car_brand']),str(goods_car_data['car_series']),str(goods_car_data['car_model']),
                                                         str(goods_car_data['car_power']),str(goods_car_data['car_year']),str(goods_car_data['car_name'])))
            is_sheet_num += 1

        # 不匹配
        for car_id in now_all_car_list:
            goods_car_data = self.car_category_dict[car_id]
            self.write_excel_cell(not_sheet,not_sheet_num,(str(car_id),str(goods_car_data['brand']),str(goods_car_data['series']),str(goods_car_data['model']),
                                                           str(goods_car_data['power']),str(goods_car_data['year']),str(goods_car_data['name'])))
            not_sheet_num += 1
        ew.save(filename=self.file_parent+'/sku/'+goods_id+'.xlsx')

        self.cat_car_dict[cat_id] = cat_car_list


        print "%s goods excel is ok" % goods_id

    def stat_cat(self):
        sum_wb = Workbook()
        sum_ew = ExcelWriter(workbook=sum_wb)
        sum_sheet = sum_wb.create_sheet(u"统计", 0)
        sum_sheet.cell(row=1, column=1).value = u'分类名'
        sum_sheet.cell(row=1, column=2).value = u'总品牌数'
        sum_sheet.cell(row=1, column=3).value = u'覆盖的品牌数'
        sum_sheet.cell(row=1, column=4).value = u'总最小子级车型数'
        sum_sheet.cell(row=1, column=5).value = u'覆盖数'
        sum_sheet.cell(row=1, column=6).value = u'覆盖品牌比例'
        sum_sheet.cell(row=1, column=7).value = u'覆盖最小子级车型比例'

        sum_num = 2

        for cat_id,cat_car_list in self.cat_car_dict.iteritems():
            cat_name = self.cat_dict[cat_id]
            fu_car_num = len(cat_car_list)
            fu_brand_list = list()

            wb = Workbook()
            ew = ExcelWriter(workbook=wb)
            is_sheet = wb.create_sheet(u"适配车型", 0)
            not_sheet = wb.create_sheet(u"不适配车型", 1)
            is_sheet_num = 2
            not_sheet_num = 2
            # 初始化 sheet
            for the_sheet in (is_sheet,not_sheet):
                insert_list = (u'car_id',u'品牌',u'车系',u'车型',u'排量',u'年款',u'名称')
                self.write_excel_cell(the_sheet,1,insert_list)

            now_all_car_list = copy.deepcopy(self.car_category_list)
            for car_id in cat_car_list:
                now_all_car_list.remove(car_id)
                goods_car_data = self.car_category_dict[car_id]
                brand = str(goods_car_data['brand'])
                if brand not in fu_brand_list:
                    fu_brand_list.append(brand)
                self.write_excel_cell(is_sheet,is_sheet_num,(car_id,brand,str(goods_car_data['series']),str(goods_car_data['model']),
                                                               str(goods_car_data['power']),str(goods_car_data['year']),str(goods_car_data['name'])))
                is_sheet_num += 1

            # 不匹配
            for car_id in now_all_car_list:
                goods_car_data = self.car_category_dict[car_id]
                self.write_excel_cell(not_sheet,not_sheet_num,(car_id,str(goods_car_data['brand']),str(goods_car_data['series']),str(goods_car_data['model']),
                                                               str(goods_car_data['power']),str(goods_car_data['year']),str(goods_car_data['name'])))
                not_sheet_num += 1
            ew.save(filename=self.file_parent+'/'+cat_name+'.xlsx')
            fu_brand_num = len(fu_brand_list)

            self.write_excel_cell(sum_sheet,sum_num,(cat_name,str(self.brand_sum_num),str(fu_brand_num),str(self.car_sum_num),str(fu_car_num),str(fu_brand_num/self.brand_sum_num),str(fu_car_num/self.car_sum_num)))
            sum_num +=1

        sum_ew.save(filename=self.file_parent+'/总的统计.xlsx')


    def start(self):
        print "=====start====="
        goods_sql = "select goods_id,cat_id,goods_name from db_goods where brand_id = "+str(self.brand_id)+" and is_delete = 0"
        # goods_sql = "select goods_id,cat_id,goods_name from db_goods where cat_id = 5025 and is_delete = 0"
        goods_array = self.shop_dao.db.get_data(goods_sql)
        for goods_data in goods_array:
            goods_id = str(goods_data['goods_id'])
            cat_id = str(goods_data['cat_id'])
            goods_name = str(goods_data['goods_name'])

            self.stat_sku(goods_id,goods_name,cat_id)

        # 分类 excle 出
        self.stat_cat()

        print "=====end====="



goodsCarPercentage = GoodsCarPercentage()
goodsCarPercentage.start()